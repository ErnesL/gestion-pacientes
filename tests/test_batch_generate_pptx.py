from __future__ import annotations

import io
import sys
import unittest
from contextlib import redirect_stdout
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from batch_generate_pptx import generate_pptx_batch  # noqa: E402


SUMMARY_ROWS = [
    ("Evaluación", "1era evaluación"),
    ("Fecha", datetime(2026, 3, 25)),
    ("Peso (Kg)", 73.1),
    ("Talla parada (cm)", 173.5),
    ("% Grasa (Carter 1986)", 26.0263),
    ("Kg de Grasa", 19.0252253),
]

MEASUREMENT_ROWS = [
    ("Fecha de evaluación", datetime(2026, 3, 25)),
    ("Peso actual (kg)", 73.1),
    ("Talla (m)", 1.735),
    ("Talla (cm)", 173.5),
]


def build_valid_workbook(path: Path, *, patient_name: str) -> None:
    wb = Workbook()

    history = wb.active
    history.title = "HISTORIA"
    history["C4"] = patient_name
    history["C5"] = "12345678"
    history["C7"] = 23
    history["C10"] = "Femenino"
    history["I8"] = "Fitness"

    plan = wb.create_sheet("PLAN_ALIMENTACION_TEMPLATE")
    plan.append(
        [
            "COMIDA",
            "LACTEOS",
            "VEGETALES",
            "FRUTAS",
            "ALMIDONES",
            "PROTEINAS",
            "GRASAS",
        ]
    )
    plan.append(["DES", 1, 0, 1, 2, 1, 1])

    anthro = wb.create_sheet("ANTROPOMETRIA_TEMPLATE")
    anthro.append(["SECCION", "ETIQUETA", "VALOR"])
    for label, value in SUMMARY_ROWS:
        anthro.append(["RESUMEN", label, value])
    for label, value in MEASUREMENT_ROWS:
        anthro.append(["MEDIDAS", label, value])

    wb.save(path)


class BatchGeneratePptxTest(unittest.TestCase):
    def test_generate_pptx_batch_outputs_next_to_each_excel_and_writes_reports(self) -> None:
        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            first_dir = root / "clientes" / "Paciente Uno"
            second_dir = root / "clientes" / "Paciente Dos"
            first_dir.mkdir(parents=True)
            second_dir.mkdir(parents=True)

            first_excel = first_dir / "historia-uno.xlsx"
            second_excel = second_dir / "historia-dos.xlsx"
            build_valid_workbook(first_excel, patient_name="Paciente Uno")
            build_valid_workbook(second_excel, patient_name="Paciente Dos")

            buffer = io.StringIO()
            with redirect_stdout(buffer):
                results = generate_pptx_batch(root, show_progress=True)

            self.assertEqual(len(results), 2)
            self.assertTrue((first_dir / "Plan Alimentacion - Paciente Uno.pptx").exists())
            self.assertTrue((first_dir / "Informe Antropometrico - Paciente Uno.pptx").exists())
            self.assertTrue((second_dir / "Plan Alimentacion - Paciente Dos.pptx").exists())
            self.assertTrue((second_dir / "Informe Antropometrico - Paciente Dos.pptx").exists())
            self.assertTrue((root / "reporte-generacion-pptx.csv").exists())
            self.assertTrue((root / "resumen-generacion-pptx.txt").exists())

            output = buffer.getvalue()
            self.assertIn("Voy por: clientes/Paciente Dos/historia-dos.xlsx | quedan 1", output)
            self.assertIn("Voy por: clientes/Paciente Uno/historia-uno.xlsx | quedan 0", output)
            self.assertIn("-> OK", output)


if __name__ == "__main__":
    unittest.main()
