from __future__ import annotations

import sys
import unittest
from datetime import datetime, time
from pathlib import Path
from tempfile import TemporaryDirectory
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from add_template_sheets import (  # noqa: E402
    DEFAULT_TEMPLATE_WORKBOOK,
    selected_sheet_names,
    update_folder_with_template_sheets,
    update_workbook_with_template_sheets,
)


def build_base_workbook(path: Path, *, include_plan: bool = False) -> None:
    wb = Workbook()
    history = wb.active
    history.title = "HISTORIA"
    history["C4"] = "Paciente Demo"
    history["C5"] = "12345678"
    history["C7"] = 30
    history["C10"] = "Femenino"
    history["I8"] = "Fitness"

    if include_plan:
        plan = wb.create_sheet("PLAN_ALIMENTACION_TEMPLATE")
        plan.append(["COMIDA", "LACTEOS", "VEGETALES", "FRUTAS",
                    "ALMIDONES", "PROTEINAS", "GRASAS"])
        plan.append(["DES", 9, 9, 9, 9, 9, 9])

    wb.save(path)


def build_requerimientos_plan_workbook(path: Path) -> None:
    wb = Workbook()
    history = wb.active
    history.title = "HISTORIA"
    history["C4"] = "Paciente Plan"
    history["C5"] = "22223333"
    history["C7"] = 29
    history["C10"] = "Masculino"

    requirements = wb.create_sheet("REQUERIMIENTOS")
    requirements["J47"] = "Comida"
    requirements["K47"] = "Pre Desayuno"
    requirements["L47"] = "Desayuno"
    requirements["M47"] = "Merienda"
    requirements["N47"] = "Almuerzo"
    requirements["P47"] = "Merienda"
    requirements["R47"] = "Cena"
    requirements["J48"] = "Leche"
    requirements["J49"] = "Vegetales"
    requirements["J50"] = "Frutas"
    requirements["J51"] = "Almidon"
    requirements["J53"] = "Carnes semi"
    requirements["J54"] = "Grasas"
    requirements["L50"] = 1
    requirements["N49"] = 2
    requirements["N50"] = 1
    requirements["L51"] = 4
    requirements["N51"] = 5
    requirements["L53"] = 4
    requirements["N53"] = 5
    requirements["L54"] = 3
    requirements["N54"] = 3

    wb.save(path)


def build_existing_anthro_template_workbook(path: Path) -> None:
    wb = Workbook()
    history = wb.active
    history.title = "HISTORIA"
    history["C4"] = "Paciente Demo"
    history["C5"] = "12345678"
    history["C7"] = 30
    history["C10"] = "Femenino"

    anthro = wb.create_sheet("ANTROPOMETRIA_TEMPLATE")
    anthro.append(["SECCION", "ETIQUETA", "VALOR", "CONTROL_1"])
    anthro.append(["RESUMEN", "Evaluación", "1era evaluación", "Control 1°"])
    anthro.append(["RESUMEN", "Fecha", "01/01/2026", "01/02/2026"])
    anthro.append(["RESUMEN", "Peso (Kg)", 71.4, 69.9])
    anthro.append(["RESUMEN", "Talla parada (cm)", 173.0, 173.0])
    anthro.append(["RESUMEN", "% Grasa (Carter 1986)", 24.1, 22.4])
    anthro.append(["RESUMEN", "Kg de Grasa", 17.2, 15.7])
    anthro.append(["MEDIDAS", "Fecha de evaluación", "01/01/2026", "01/02/2026"])
    anthro.append(["MEDIDAS", "Talla (m)", 1.73, 1.73])
    wb.save(path)


def build_legacy_summary_workbook(path: Path) -> None:
    wb = Workbook()
    history = wb.active
    history.title = "HISTORIA"
    history["C4"] = "Paciente Legacy"
    history["C5"] = "87654321"
    history["C7"] = 28
    history["C10"] = "Masculino"

    legacy = wb.create_sheet("RESUMEN ANTROPOMETRICO")
    legacy["B2"] = "Evaluación"
    legacy["D2"] = "1era evaluación"
    legacy["E2"] = "Control 1°"
    legacy["B3"] = "Fecha de evaluación"
    legacy["D3"] = "03/01/2026"
    legacy["E3"] = "03/02/2026"
    legacy["B4"] = "Peso Actual (Kg)"
    legacy["D4"] = 80.2
    legacy["E4"] = 78.6
    legacy["B5"] = "Talla parada (cm)"
    legacy["D5"] = 180
    legacy["E5"] = 180
    legacy["B6"] = "% Grasa (Carter 1986)"
    legacy["D6"] = 22.1
    legacy["E6"] = 20.4
    legacy["B9"] = "Kg de Masa Magra"
    legacy["D9"] = 62.5
    legacy["E9"] = 62.6
    legacy["B10"] = "Kg de Grasa"
    legacy["D10"] = 17.7
    legacy["E10"] = 16.0
    legacy["B13"] = "Sumatoria de 6 pliegues"
    legacy["D13"] = 120
    legacy["E13"] = 111
    legacy["B14"] = "Somatotipo"
    legacy["D14"] = "Mesomorfo"
    legacy["E14"] = "Mesomorfo"
    legacy["B19"] = "Fecha de evaluación"
    legacy["D19"] = "03/01/2026"
    legacy["E19"] = "03/02/2026"
    legacy["B20"] = "Peso actual (kg)"
    legacy["D20"] = 80.2
    legacy["E20"] = 78.6
    legacy["B21"] = "Talla (m)"
    legacy["D21"] = 1.80
    legacy["E21"] = 1.80
    legacy["B22"] = "Talla (cm)"
    legacy["D22"] = 180
    legacy["E22"] = 180
    legacy["B24"] = "Brazo relajado (cm)"
    legacy["D24"] = 33.1
    legacy["E24"] = 32.6
    wb.save(path)


def build_legacy_summary_workbook_with_invalid_trailing_columns(path: Path) -> None:
    wb = Workbook()
    history = wb.active
    history.title = "HISTORIA"
    history["C4"] = "Paciente Legacy"
    history["C5"] = "87654321"
    history["C7"] = 28
    history["C10"] = "Masculino"

    legacy = wb.create_sheet("RESUMEN ANTROPOMETRICO")
    legacy["B2"] = "Evaluación"
    legacy["D2"] = "1era evaluación"
    legacy["E2"] = "Control 1°"
    legacy["F2"] = "Control 2°"
    legacy["B3"] = "Fecha de evaluación"
    legacy["D3"] = datetime(2026, 1, 3)
    legacy["E3"] = datetime(2026, 2, 3)
    legacy["F3"] = time(0, 0)
    legacy["B4"] = "Peso Actual (Kg)"
    legacy["D4"] = 80.2
    legacy["E4"] = 78.6
    legacy["F4"] = 0
    legacy["B5"] = "Talla parada (cm)"
    legacy["D5"] = 180
    legacy["E5"] = 180
    legacy["F5"] = 0
    legacy["B6"] = "% Grasa (Carter 1986)"
    legacy["D6"] = 22.1
    legacy["E6"] = 20.4
    legacy["F6"] = 3.5
    legacy["B10"] = "Kg de Grasa"
    legacy["D10"] = 17.7
    legacy["E10"] = 16.0
    legacy["F10"] = 0
    legacy["B19"] = "Fecha de evaluación"
    legacy["D19"] = datetime(2026, 1, 3)
    legacy["E19"] = datetime(2026, 2, 3)
    legacy["F19"] = time(0, 0)
    legacy["B20"] = "Peso actual (kg)"
    legacy["D20"] = 80.2
    legacy["E20"] = 78.6
    legacy["F20"] = 0
    legacy["B21"] = "Talla (m)"
    legacy["D21"] = 1.80
    legacy["E21"] = 1.80
    legacy["F21"] = 0
    wb.save(path)


def build_aaron_style_summary_workbook(path: Path) -> None:
    wb = Workbook()
    history = wb.active
    history.title = "HISTORIA"
    history["C4"] = "Aaron Lopez"
    history["C5"] = "12345678"
    history["C10"] = "Masculino"

    legacy = wb.create_sheet("RESUMEN ANTROPOMETRICO")
    legacy["D16"] = "Evaluación"
    legacy["F16"] = 1
    legacy["G16"] = 2
    legacy["D17"] = "Fecha"
    legacy["F17"] = datetime(2025, 12, 1)
    legacy["G17"] = time(0, 0)
    legacy["D18"] = "Peso (Kg)"
    legacy["F18"] = 62.35
    legacy["G18"] = 0
    legacy["D19"] = "Talla parada (cm)"
    legacy["F19"] = 175
    legacy["G19"] = 185
    legacy["D20"] = "% Grasa (Carter 1986)"
    legacy["F20"] = 11.04555
    legacy["G20"] = 2.585
    legacy["D22"] = "Kg de Grasa"
    legacy["F22"] = 6.8869
    legacy["G22"] = 0
    legacy["E34"] = "Fecha de evaluación"
    legacy["F34"] = time(0, 0)
    legacy["G34"] = time(0, 0)
    legacy["E35"] = "Peso actual (kg)"
    legacy["F35"] = 0
    legacy["G35"] = 0
    legacy["E36"] = "Talla (m)"
    legacy["F36"] = 0
    legacy["G36"] = 0
    legacy["E37"] = "Talla (cm)"
    legacy["F37"] = 0
    legacy["G37"] = 0
    wb.save(path)


def build_formula_driven_summary_workbook(path: Path) -> None:
    wb = Workbook()
    history = wb.active
    history.title = "HISTORIA"
    history["C4"] = "Aaron Lopez"
    history["C5"] = "12345678"
    history["C10"] = "Masculino"

    body = wb.create_sheet("5 COMPONENTES")
    body["C2"] = datetime(2025, 12, 1)
    body["C5"] = 62.35
    body["C6"] = 175
    body["C10"] = 7
    body["C11"] = 8.5
    body["C14"] = 18
    body["C15"] = 29.5
    body["C16"] = 11
    body["C17"] = 6.5
    body["C70"] = "=C10+C11+C15+C14+C16+C17"
    body["C80"] = "=(2.585)+(0.1051*(C10+C11+C15+C14+C16+C17))"
    body["C81"] = "=C80"
    body["C82"] = "=((100-C81)*C5)/100"
    body["C83"] = "=C5-C82"
    body["C111"] = "Ectomorfo"

    legacy = wb.create_sheet("RESUMEN ANTROPOMETRICO")
    legacy["D16"] = "Evaluación"
    legacy["F16"] = 1
    legacy["D17"] = "Fecha"
    legacy["F17"] = "='5 COMPONENTES'!C2"
    legacy["D18"] = "Peso (Kg)"
    legacy["F18"] = "='5 COMPONENTES'!C5"
    legacy["D19"] = "Talla parada (cm)"
    legacy["F19"] = "='5 COMPONENTES'!C6"
    legacy["D20"] = "% Grasa (Carter 1986)"
    legacy["F20"] = "='5 COMPONENTES'!C80"
    legacy["D21"] = "Kg de Masa Magra"
    legacy["F21"] = "='5 COMPONENTES'!C82"
    legacy["D22"] = "Kg de Grasa"
    legacy["F22"] = "='5 COMPONENTES'!C83"
    legacy["D27"] = "Masa Adiposa (Kg)"
    legacy["F27"] = "='5 COMPONENTES'!C83"
    legacy["D28"] = "Sumatoria de 6 pliegues"
    legacy["F28"] = "='5 COMPONENTES'!C70"
    legacy["D29"] = "Somatotipo"
    legacy["F29"] = "='5 COMPONENTES'!C111"

    legacy["E34"] = "Fecha de evaluación"
    legacy["F34"] = "='5 COMPONENTES'!C2"
    legacy["E35"] = "Peso actual (kg)"
    legacy["F35"] = "='5 COMPONENTES'!C5"
    legacy["E36"] = "Talla (m)"
    legacy["F36"] = "='5 COMPONENTES'!C6/100"
    legacy["E37"] = "Talla (cm)"
    legacy["F37"] = "='5 COMPONENTES'!C6"
    wb.save(path)


def build_workbook_without_anthro_bases(path: Path) -> None:
    wb = Workbook()
    history = wb.active
    history.title = "HISTORIA"
    history["C4"] = "Paciente Sin Base"
    history["C5"] = "11112222"
    history["C7"] = 25
    history["C10"] = "Femenino"
    wb.save(path)


class AddTemplateSheetsTest(unittest.TestCase):
    def assert_workbook_has_no_overlapping_column_ranges(self, workbook_path: Path) -> None:
        namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        with ZipFile(workbook_path) as archive:
            for member in archive.namelist():
                if not member.startswith("xl/worksheets/sheet") or not member.endswith(".xml"):
                    continue
                root = ElementTree.fromstring(archive.read(member))
                cols = root.find("main:cols", namespace)
                if cols is None:
                    continue
                seen_ranges: list[tuple[int, int]] = []
                for col in cols.findall("main:col", namespace):
                    min_col = int(col.attrib["min"])
                    max_col = int(col.attrib["max"])
                    for previous_min, previous_max in seen_ranges:
                        overlaps = not (
                            max_col < previous_min or min_col > previous_max
                        )
                        self.assertFalse(
                            overlaps,
                            f"{member} has overlapping column ranges "
                            f"{(min_col, max_col)} and {(previous_min, previous_max)}",
                        )
                    seen_ranges.append((min_col, max_col))

    def test_batch_adds_default_sheets_without_copying_dummy_plan_examples(self) -> None:
        with TemporaryDirectory() as tmpdir:
            temp_dir = Path(tmpdir)
            excel_path = temp_dir / "paciente-1.xlsx"
            build_base_workbook(excel_path)

            results = update_folder_with_template_sheets(
                temp_dir,
                template_workbook_path=DEFAULT_TEMPLATE_WORKBOOK,
                sheet_names=selected_sheet_names(),
            )

            self.assertEqual(len(results), 1)
            self.assertEqual(
                results[0].added_sheets,
                (
                    "PLAN_ALIMENTACION_TEMPLATE",
                    "ANTROPOMETRIA_TEMPLATE",
                    "EJEMPLOS_COMIDAS",
                    "EQUIVALENCIAS_EJEMPLOS",
                ),
            )
            self.assertEqual(results[0].replaced_sheets, ())
            self.assertEqual(results[0].skipped_sheets, ())

            workbook = load_workbook(excel_path)
            self.assertIn("PLAN_ALIMENTACION_TEMPLATE", workbook.sheetnames)
            self.assertIn("ANTROPOMETRIA_TEMPLATE", workbook.sheetnames)
            self.assertIn("EJEMPLOS_COMIDAS", workbook.sheetnames)
            self.assertIn("EQUIVALENCIAS_EJEMPLOS", workbook.sheetnames)
            self.assertEqual(
                workbook["PLAN_ALIMENTACION_TEMPLATE"]["A2"].value,
                None,
            )
            self.assertEqual(
                workbook["ANTROPOMETRIA_TEMPLATE"]["B2"].value,
                "Evaluación",
            )
            self.assertEqual(
                workbook["ANTROPOMETRIA_TEMPLATE"]["C1"].value,
                "VALOR",
            )
            self.assertEqual(
                workbook["ANTROPOMETRIA_TEMPLATE"]["F1"].value,
                "CONTROL_3",
            )
            self.assertIsNone(workbook["ANTROPOMETRIA_TEMPLATE"]["F2"].value)
            self.assertEqual(
                workbook["EJEMPLOS_COMIDAS"]["H2"].value,
                None,
            )
            self.assertEqual(workbook["EJEMPLOS_COMIDAS"]["A2"].value, "PRE")
            self.assertEqual(workbook["EJEMPLOS_COMIDAS"]["B2"].value, "yogurt griego")
            self.assertEqual(workbook["EJEMPLOS_COMIDAS"]["D2"].value, "cambur")
            self.assertEqual(workbook["EJEMPLOS_COMIDAS"]["F2"].value, "proteina liquida")
            self.assertEqual(workbook["EJEMPLOS_COMIDAS"].row_dimensions[3].height, 18.0)
            self.assertTrue(workbook["EJEMPLOS_COMIDAS"].column_dimensions["I"].hidden)
            self.assert_workbook_has_no_overlapping_column_ranges(excel_path)

    def test_batch_preserves_existing_plan_sheet_data(self) -> None:
        with TemporaryDirectory() as tmpdir:
            temp_dir = Path(tmpdir)
            excel_path = temp_dir / "paciente-2.xlsx"
            build_base_workbook(excel_path, include_plan=True)

            results = update_folder_with_template_sheets(
                temp_dir,
                template_workbook_path=DEFAULT_TEMPLATE_WORKBOOK,
                sheet_names=selected_sheet_names(),
            )

            self.assertEqual(len(results), 1)
            self.assertEqual(
                results[0].added_sheets,
                (
                    "ANTROPOMETRIA_TEMPLATE",
                    "EJEMPLOS_COMIDAS",
                    "EQUIVALENCIAS_EJEMPLOS",
                ),
            )
            self.assertEqual(
                results[0].replaced_sheets,
                ("PLAN_ALIMENTACION_TEMPLATE",),
            )
            self.assertEqual(results[0].skipped_sheets, ())

            workbook = load_workbook(excel_path)
            self.assertEqual(
                workbook["PLAN_ALIMENTACION_TEMPLATE"]["B2"].value,
                9,
            )
            self.assertIn("EQUIVALENCIAS_EJEMPLOS", workbook.sheetnames)

    def test_replace_existing_template_preserves_existing_anthro_data(self) -> None:
        with TemporaryDirectory() as tmpdir:
            temp_dir = Path(tmpdir)
            excel_path = temp_dir / "paciente-3.xlsx"
            build_existing_anthro_template_workbook(excel_path)

            result = update_workbook_with_template_sheets(
                excel_path,
                template_workbook_path=DEFAULT_TEMPLATE_WORKBOOK,
                sheet_names=selected_sheet_names(),
                replace_existing=True,
            )

            self.assertEqual(result.replaced_sheets, ("ANTROPOMETRIA_TEMPLATE",))
            workbook = load_workbook(excel_path)
            anthro = workbook["ANTROPOMETRIA_TEMPLATE"]
            self.assertEqual(anthro["C1"].value, "VALOR")
            self.assertEqual(anthro["D1"].value, "CONTROL_1")
            self.assertEqual(anthro["C2"].value, "1era evaluación")
            self.assertEqual(anthro["D2"].value, "Control 1°")
            label_to_row = {
                anthro.cell(row=row_idx, column=2).value: row_idx
                for row_idx in range(2, anthro.max_row + 1)
            }
            peso_row = label_to_row["Peso (Kg)"]
            self.assertEqual(anthro.cell(row=peso_row, column=3).value, 71.4)
            self.assertEqual(anthro.cell(row=peso_row, column=4).value, 69.9)

    def test_adding_template_backfills_from_legacy_summary_sheet(self) -> None:
        with TemporaryDirectory() as tmpdir:
            temp_dir = Path(tmpdir)
            excel_path = temp_dir / "paciente-4.xlsx"
            build_legacy_summary_workbook(excel_path)

            result = update_workbook_with_template_sheets(
                excel_path,
                template_workbook_path=DEFAULT_TEMPLATE_WORKBOOK,
                sheet_names=selected_sheet_names(),
            )

            self.assertEqual(
                result.added_sheets,
                (
                    "PLAN_ALIMENTACION_TEMPLATE",
                    "ANTROPOMETRIA_TEMPLATE",
                    "EJEMPLOS_COMIDAS",
                    "EQUIVALENCIAS_EJEMPLOS",
                ),
            )
            workbook = load_workbook(excel_path)
            anthro = workbook["ANTROPOMETRIA_TEMPLATE"]
            self.assertEqual(anthro["C1"].value, "VALOR")
            self.assertEqual(anthro["D1"].value, "CONTROL_1")
            self.assertEqual(anthro["C2"].value, "1era evaluación")
            self.assertEqual(anthro["D2"].value, "Control 1°")
            self.assertEqual(anthro["B3"].value, "Fecha")
            self.assertEqual(anthro["C3"].value, "03/01/2026")
            self.assertEqual(anthro["D3"].value, "03/02/2026")
            self.assertEqual(anthro["B4"].value, "Peso (Kg)")
            self.assertEqual(anthro["C4"].value, 80.2)
            self.assertEqual(anthro["D4"].value, 78.6)

    def test_adding_template_without_valid_bases_leaves_anthro_values_blank(self) -> None:
        with TemporaryDirectory() as tmpdir:
            temp_dir = Path(tmpdir)
            excel_path = temp_dir / "paciente-5.xlsx"
            build_workbook_without_anthro_bases(excel_path)

            result = update_workbook_with_template_sheets(
                excel_path,
                template_workbook_path=DEFAULT_TEMPLATE_WORKBOOK,
                sheet_names=selected_sheet_names(),
            )

            self.assertEqual(
                result.added_sheets,
                (
                    "PLAN_ALIMENTACION_TEMPLATE",
                    "ANTROPOMETRIA_TEMPLATE",
                    "EJEMPLOS_COMIDAS",
                    "EQUIVALENCIAS_EJEMPLOS",
                ),
            )
            workbook = load_workbook(excel_path)
            anthro = workbook["ANTROPOMETRIA_TEMPLATE"]
            self.assertEqual(anthro["C1"].value, "VALOR")
            self.assertEqual(anthro["F1"].value, "CONTROL_3")
            self.assertIsNone(anthro["C2"].value)
            self.assertIsNone(anthro["F2"].value)
            self.assertIsNone(anthro["C10"].value)
            self.assertIsNone(workbook["PLAN_ALIMENTACION_TEMPLATE"]["A2"].value)
            self.assertEqual(workbook["EJEMPLOS_COMIDAS"]["A2"].value, "PRE")
            self.assertIsNone(workbook["EJEMPLOS_COMIDAS"]["H2"].value)

    def test_backfill_discards_invalid_trailing_anthro_columns(self) -> None:
        with TemporaryDirectory() as tmpdir:
            temp_dir = Path(tmpdir)
            excel_path = temp_dir / "paciente-6.xlsx"
            build_legacy_summary_workbook_with_invalid_trailing_columns(excel_path)

            update_workbook_with_template_sheets(
                excel_path,
                template_workbook_path=DEFAULT_TEMPLATE_WORKBOOK,
                sheet_names=selected_sheet_names(),
            )

            workbook = load_workbook(excel_path)
            anthro = workbook["ANTROPOMETRIA_TEMPLATE"]
            self.assertEqual(anthro["C1"].value, "VALOR")
            self.assertEqual(anthro["D1"].value, "CONTROL_1")
            self.assertIsNone(anthro["E1"].value)
            self.assertEqual(anthro["C2"].value, "1era evaluación")
            self.assertEqual(anthro["D2"].value, "Control 1°")
            self.assertIsNone(anthro["E2"].value)
            self.assertEqual(anthro["C4"].value, 80.2)
            self.assertEqual(anthro["D4"].value, 78.6)

    def test_backfill_supports_aaron_summary_layout(self) -> None:
        with TemporaryDirectory() as tmpdir:
            temp_dir = Path(tmpdir)
            excel_path = temp_dir / "paciente-7.xlsx"
            build_aaron_style_summary_workbook(excel_path)

            update_workbook_with_template_sheets(
                excel_path,
                template_workbook_path=DEFAULT_TEMPLATE_WORKBOOK,
                sheet_names=selected_sheet_names(),
            )

            workbook = load_workbook(excel_path)
            anthro = workbook["ANTROPOMETRIA_TEMPLATE"]
            label_to_row = {
                anthro.cell(row=row_idx, column=2).value: row_idx
                for row_idx in range(2, anthro.max_row + 1)
            }
            self.assertEqual(anthro["C1"].value, "VALOR")
            self.assertIsNone(anthro["D1"].value)
            self.assertEqual(anthro["C2"].value, "1era evaluación")
            self.assertEqual(
                anthro.cell(row=label_to_row["Fecha"], column=3).value,
                datetime(2025, 12, 1),
            )
            self.assertEqual(
                anthro.cell(row=label_to_row["Peso (Kg)"], column=3).value,
                62.35,
            )
            self.assertEqual(
                anthro.cell(row=label_to_row["Talla (m)"], column=3).value,
                1.75,
            )

    def test_backfill_resolves_formula_driven_summary_without_cached_values(self) -> None:
        with TemporaryDirectory() as tmpdir:
            temp_dir = Path(tmpdir)
            excel_path = temp_dir / "paciente-7b.xlsx"
            build_formula_driven_summary_workbook(excel_path)

            update_workbook_with_template_sheets(
                excel_path,
                template_workbook_path=DEFAULT_TEMPLATE_WORKBOOK,
                sheet_names=selected_sheet_names(),
            )

            workbook = load_workbook(excel_path)
            anthro = workbook["ANTROPOMETRIA_TEMPLATE"]
            label_to_row = {
                anthro.cell(row=row_idx, column=2).value: row_idx
                for row_idx in range(2, anthro.max_row + 1)
            }
            self.assertEqual(anthro["C2"].value, "1era evaluación")
            self.assertEqual(
                anthro.cell(row=label_to_row["Fecha"], column=3).value,
                datetime(2025, 12, 1),
            )
            self.assertEqual(
                anthro.cell(row=label_to_row["Peso (Kg)"], column=3).value,
                62.35,
            )
            self.assertAlmostEqual(
                anthro.cell(row=label_to_row["% Grasa (Carter 1986)"], column=3).value,
                11.04555,
                places=5,
            )
            self.assertAlmostEqual(
                anthro.cell(row=label_to_row["Kg de Grasa"], column=3).value,
                6.886900425,
                places=6,
            )
            self.assertEqual(
                anthro.cell(row=label_to_row["Peso actual (kg)"], column=3).value,
                62.35,
            )
            self.assertEqual(
                anthro.cell(row=label_to_row["Talla (m)"], column=3).value,
                1.75,
            )

    def test_backfill_plan_template_from_requerimientos_sheet(self) -> None:
        with TemporaryDirectory() as tmpdir:
            temp_dir = Path(tmpdir)
            excel_path = temp_dir / "paciente-8.xlsx"
            build_requerimientos_plan_workbook(excel_path)

            update_workbook_with_template_sheets(
                excel_path,
                template_workbook_path=DEFAULT_TEMPLATE_WORKBOOK,
                sheet_names=selected_sheet_names(),
            )

            workbook = load_workbook(excel_path)
            plan = workbook["PLAN_ALIMENTACION_TEMPLATE"]
            self.assertEqual(plan["A2"].value, "PRE")
            self.assertEqual(plan["A3"].value, "DES")
            self.assertEqual(plan["B3"].value, 0)
            self.assertEqual(plan["C3"].value, 0)
            self.assertEqual(plan["D3"].value, 1)
            self.assertEqual(plan["E3"].value, 4)
            self.assertEqual(plan["F3"].value, 4)
            self.assertEqual(plan["G3"].value, 3)
            self.assertEqual(plan["A5"].value, "ALM")
            self.assertEqual(plan["C5"].value, 2)
            self.assertEqual(plan["D5"].value, 1)
            self.assertEqual(plan["E5"].value, 5)
            self.assertEqual(plan["F5"].value, 5)
            self.assertEqual(plan["G5"].value, 3)


if __name__ == "__main__":
    unittest.main()
