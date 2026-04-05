from __future__ import annotations

import io
import sys
import unittest
from contextlib import redirect_stdout
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from filter_recent_excels import (  # noqa: E402
    copy_selected_workbooks,
    detect_reference_date,
    load_excel_workbook,
    parse_date_like,
)


def build_history_date_workbook(path: Path, *, history_date: datetime) -> None:
    wb = Workbook()
    history = wb.active
    history.title = "HISTORIA"
    history["B2"] = history_date
    history["C4"] = "Paciente Demo"
    history["C5"] = "12345678"
    history["C10"] = "Femenino"
    wb.save(path)


def build_legacy_anthro_workbook(path: Path, *, evaluation_date: datetime) -> None:
    wb = Workbook()
    history = wb.active
    history.title = "HISTORIA"
    history["C4"] = "Paciente Demo"
    history["C5"] = "12345678"
    history["C10"] = "Masculino"

    legacy = wb.create_sheet("RESUMEN ANTROPOMETRICO")
    legacy["B2"] = "Evaluación"
    legacy["D2"] = "1era evaluación"
    legacy["B3"] = "Fecha de evaluación"
    legacy["D3"] = evaluation_date
    legacy["B4"] = "Peso Actual (Kg)"
    legacy["D4"] = 80.2
    legacy["B5"] = "Talla parada (cm)"
    legacy["D5"] = 180
    legacy["B6"] = "% Grasa (Carter 1986)"
    legacy["D6"] = 20.4
    legacy["B10"] = "Kg de Grasa"
    legacy["D10"] = 16.0
    legacy["B19"] = "Fecha de evaluación"
    legacy["D19"] = evaluation_date
    legacy["B20"] = "Peso actual (kg)"
    legacy["D20"] = 80.2
    legacy["B21"] = "Talla (m)"
    legacy["D21"] = 1.80
    wb.save(path)


class FilterRecentExcelsTest(unittest.TestCase):
    def test_parse_date_like_supports_common_formats(self) -> None:
        self.assertEqual(parse_date_like("2025-03-14").isoformat(), "2025-03-14")
        self.assertEqual(parse_date_like("14/03/2025").isoformat(), "2025-03-14")
        self.assertEqual(parse_date_like("14/03/25").isoformat(), "2025-03-14")

    def test_detect_reference_date_prefers_history_b2(self) -> None:
        with TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "paciente.xlsx"
            build_history_date_workbook(
                workbook_path,
                history_date=datetime(2025, 3, 14),
            )

            value_wb = load_excel_workbook(workbook_path, data_only=True)
            formula_wb = load_excel_workbook(workbook_path, data_only=False)
            try:
                selected_date, source = detect_reference_date(
                    workbook_path,
                    value_wb,
                    formula_wb,
                )
            finally:
                formula_wb.close()
                value_wb.close()

            self.assertEqual(selected_date.isoformat(), "2025-03-14")
            self.assertEqual(source, "HISTORIA!B2")

    def test_copy_selected_workbooks_preserves_organization_tree(self) -> None:
        with TemporaryDirectory() as tmpdir:
            base_dir = Path(tmpdir)
            input_root = base_dir / "excels"
            output_root = base_dir / "salida"
            caracas_dir = input_root / "CARACAS" / "Pacientes" / "Paciente Uno"
            element_dir = input_root / "ELEMENT BOX" / "Orcas Los Salias" / "Paciente Dos"
            caracas_dir.mkdir(parents=True)
            element_dir.mkdir(parents=True)

            recent_path = caracas_dir / "historia-reciente.xlsx"
            old_path = element_dir / "historia-vieja.xlsx"
            build_history_date_workbook(
                recent_path,
                history_date=datetime(2025, 5, 1),
            )
            build_legacy_anthro_workbook(
                old_path,
                evaluation_date=datetime(2024, 6, 15),
            )

            results = copy_selected_workbooks(
                input_root,
                output_root,
                min_year=2025,
            )

            matched = [result for result in results if result.matched]
            self.assertEqual(len(results), 2)
            self.assertEqual(len(matched), 1)
            self.assertEqual(matched[0].organization, "CARACAS")
            self.assertTrue(matched[0].backfill_applied)
            self.assertIn("PLAN_ALIMENTACION_TEMPLATE", matched[0].added_sheets)
            self.assertIn("ANTROPOMETRIA_TEMPLATE", matched[0].added_sheets)
            self.assertTrue(
                (output_root / "CARACAS" / "Pacientes" / "Paciente Uno" / "historia-reciente.xlsx").exists()
            )
            self.assertFalse(
                (output_root / "ELEMENT BOX" / "Orcas Los Salias" / "Paciente Dos" / "historia-vieja.xlsx").exists()
            )
            copied_wb = load_excel_workbook(
                output_root / "CARACAS" / "Pacientes" / "Paciente Uno" / "historia-reciente.xlsx"
            )
            try:
                self.assertIn("PLAN_ALIMENTACION_TEMPLATE", copied_wb.sheetnames)
                self.assertIn("ANTROPOMETRIA_TEMPLATE", copied_wb.sheetnames)
                self.assertIn("EJEMPLOS_COMIDAS", copied_wb.sheetnames)
                self.assertIn("EQUIVALENCIAS_EJEMPLOS", copied_wb.sheetnames)
            finally:
                copied_wb.close()

    def test_copy_selected_workbooks_can_emit_progress(self) -> None:
        with TemporaryDirectory() as tmpdir:
            base_dir = Path(tmpdir)
            input_root = base_dir / "excels"
            output_root = base_dir / "salida"
            caracas_dir = input_root / "CARACAS" / "Pacientes" / "Paciente Uno"
            element_dir = input_root / "ELEMENT BOX" / "Orcas Los Salias" / "Paciente Dos"
            caracas_dir.mkdir(parents=True)
            element_dir.mkdir(parents=True)

            recent_path = caracas_dir / "historia-reciente.xlsx"
            old_path = element_dir / "historia-vieja.xlsx"
            build_history_date_workbook(
                recent_path,
                history_date=datetime(2025, 5, 1),
            )
            build_legacy_anthro_workbook(
                old_path,
                evaluation_date=datetime(2024, 6, 15),
            )

            buffer = io.StringIO()
            with redirect_stdout(buffer):
                copy_selected_workbooks(
                    input_root,
                    output_root,
                    min_year=2025,
                    show_progress=True,
                )

            output = buffer.getvalue()
            self.assertIn("Voy por: CARACAS/Pacientes/Paciente Uno/historia-reciente.xlsx | quedan 1", output)
            self.assertIn("-> COPIADO+ACTUALIZADO | fecha 2025-05-01 | fuente HISTORIA!B2", output)
            self.assertIn("agregadas=PLAN_ALIMENTACION_TEMPLATE, ANTROPOMETRIA_TEMPLATE, EJEMPLOS_COMIDAS, EQUIVALENCIAS_EJEMPLOS", output)
            self.assertIn("Voy por: ELEMENT BOX/Orcas Los Salias/Paciente Dos/historia-vieja.xlsx | quedan 0", output)
            self.assertIn("-> OMITIDO | fecha 2024-06-15", output)


if __name__ == "__main__":
    unittest.main()
