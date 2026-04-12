from __future__ import annotations

import ast
from datetime import date, datetime, timedelta
from dataclasses import dataclass
from pathlib import Path
import re
from typing import Dict, List, Tuple

from openpyxl import load_workbook


@dataclass
class PatientInfo:
    name: str
    ci: str
    sex: str
    age: str
    discipline: str


class ValidationError(Exception):
    pass


@dataclass
class AnthropometricReportData:
    patient: PatientInfo
    peso_corporal_kg: str
    estatura_m: str
    masa_magra_kg: str
    masa_grasa_kg: str
    pct_grasa_carter: str
    table_resumen: List[List[str]]
    table_medidas: List[List[str]]


@dataclass(frozen=True)
class ExampleFood:
    code: str
    description: str
    group_code: str
    amount_per_serving: float
    use_decimal: bool
    singular_text: str
    plural_text: str


PLAN_TEMPLATE_SHEET = "PLAN_ALIMENTACION_TEMPLATE"
ANTHRO_TEMPLATE_SHEET = "ANTROPOMETRIA_TEMPLATE"

GROUP_ROWS = {
    "L": 48,  # Lacteos (Leche)
    "V": 49,  # Vegetales
    "F": 50,  # Frutas
    "A": 51,  # Almidones
    "P": 53,  # Proteinas (Carnes semi)
    "G": 54,  # Grasas
}

GROUP_SUFFIX = {
    "L": "LACTEOS",
    "V": "VEGETALES",
    "F": "FRUTAS",
    "A": "ALMIDONES",
    "P": "PROTEINAS",
    "G": "GRASAS",
}

GROUP_NAMES = {
    "L": "LACTEOS",
    "V": "VEGETALES",
    "F": "FRUTAS",
    "A": "ALMIDONES",
    "P": "PROTEINAS",
    "G": "GRASAS",
}

MEAL_DEFS = [
    {"name": "PRE", "col": "K", "groups": ["L", "V", "F", "A", "P", "G"]},
    {"name": "DES", "col": "L", "groups": ["L", "F", "A", "P", "G"]},
    {"name": "MAM", "col": "M", "groups": ["L", "F", "A", "P", "G"]},
    {"name": "ALM", "col": "N", "groups": ["V", "F", "A", "P", "G"]},
    {"name": "MTP", "col": "P", "groups": ["L", "F", "A", "P", "G"]},
    {"name": "CEN", "col": "R", "groups": ["V", "F", "A", "P", "G"]},
]

EXAMPLE_GROUP_HEADERS = {
    "LACTEOS": "L",
    "VEGETALES": "V",
    "FRUTAS": "F",
    "ALMIDONES": "A",
    "PROTEINAS": "P",
    "GRASAS": "G",
}

MEAL_EXAMPLE_ORDER = {
    "PRE": ["P", "A", "F", "L", "G", "V"],
    "DES": ["P", "A", "G", "L", "F", "V"],
    "MAM": ["L", "A", "P", "F", "G", "V"],
    "ALM": ["P", "A", "V", "G", "F", "L"],
    "MTP": ["P", "A", "L", "F", "G", "V"],
    "CEN": ["P", "A", "V", "G", "F", "L"],
}

EXAMPLE_GUIDE_VALUES = {
    "ej: lacteo",
    "ej: vegetal",
    "ej: fruta",
    "ej: almidon",
    "ej: proteina",
    "ej: grasa",
}
EXAMPLE_GUIDE_OBSERVATION = "guia: reemplazar"
INSPECTION_SHEETS = {
    "HISTORIA",
    PLAN_TEMPLATE_SHEET,
    ANTHRO_TEMPLATE_SHEET,
    "EJEMPLOS_COMIDAS",
    "EQUIVALENCIAS_EJEMPLOS",
}
FORMULA_EXACT_REFERENCE_PATTERN = re.compile(
    r"^\s*(?:(?P<sheet>'(?:[^']|'')+'|[A-Za-z0-9_ .ÁÉÍÓÚÜÑ()\-]+)!)?(?P<coord>\$?[A-Z]{1,3}\$?\d+)\s*$"
)
FORMULA_CELL_REFERENCE_PATTERN = re.compile(
    r"(?:(?P<sheet>'(?:[^']|'')+'|[A-Za-z0-9_ .ÁÉÍÓÚÜÑ()\-]+)!)?(?P<coord>\$?[A-Z]{1,3}\$?\d+)"
)

DEFAULT_EXAMPLE_FOOD_DEFS = {
    "PAN": ExampleFood(
        code="PAN",
        description="pan",
        group_code="A",
        amount_per_serving=1,
        use_decimal=False,
        singular_text="1 reb. de pan",
        plural_text="{amount} reb. de pan",
    ),
    "AREPA": ExampleFood(
        code="AREPA",
        description="arepa",
        group_code="A",
        amount_per_serving=30,
        use_decimal=False,
        singular_text="30 g de arepa",
        plural_text="{amount} g de arepa",
    ),
    "GRANOLA": ExampleFood(
        code="GRANOLA",
        description="granola",
        group_code="A",
        amount_per_serving=15,
        use_decimal=False,
        singular_text="15 g de granola",
        plural_text="{amount} g de granola",
    ),
    "ARROZ": ExampleFood(
        code="ARROZ",
        description="arroz",
        group_code="A",
        amount_per_serving=50,
        use_decimal=False,
        singular_text="50 g de arroz",
        plural_text="{amount} g de arroz",
    ),
    "PURE DE PAPA": ExampleFood(
        code="PURE DE PAPA",
        description="pure de papa",
        group_code="A",
        amount_per_serving=60,
        use_decimal=False,
        singular_text="60 g de pure de papa",
        plural_text="{amount} g de pure de papa",
    ),
    "HUEVO": ExampleFood(
        code="HUEVO",
        description="huevo",
        group_code="P",
        amount_per_serving=1,
        use_decimal=False,
        singular_text="1 huevo",
        plural_text="{amount} huevos",
    ),
    "JAMON": ExampleFood(
        code="JAMON",
        description="jamon",
        group_code="P",
        amount_per_serving=30,
        use_decimal=False,
        singular_text="30 g de jamon",
        plural_text="{amount} g de jamon",
    ),
    "POLLO": ExampleFood(
        code="POLLO",
        description="pollo",
        group_code="P",
        amount_per_serving=30,
        use_decimal=False,
        singular_text="30 g de pollo",
        plural_text="{amount} g de pollo",
    ),
    "ATUN": ExampleFood(
        code="ATUN",
        description="atun",
        group_code="P",
        amount_per_serving=30,
        use_decimal=False,
        singular_text="30 g de atun",
        plural_text="{amount} g de atun",
    ),
    "QUESO BLANCO": ExampleFood(
        code="QUESO BLANCO",
        description="queso blanco",
        group_code="P",
        amount_per_serving=30,
        use_decimal=False,
        singular_text="30 g de queso blanco",
        plural_text="{amount} g de queso blanco",
    ),
    "PROTEINA LIQUIDA": ExampleFood(
        code="PROTEINA LIQUIDA",
        description="proteina liquida",
        group_code="P",
        amount_per_serving=1,
        use_decimal=False,
        singular_text="1 servicio de proteina liquida",
        plural_text="{amount} servicios de proteina liquida",
    ),
    "YOGURT GRIEGO": ExampleFood(
        code="YOGURT GRIEGO",
        description="yogurt griego",
        group_code="L",
        amount_per_serving=170,
        use_decimal=False,
        singular_text="170 g de yogurt griego",
        plural_text="{amount} g de yogurt griego",
    ),
    "LECHE": ExampleFood(
        code="LECHE",
        description="leche",
        group_code="L",
        amount_per_serving=240,
        use_decimal=False,
        singular_text="240 ml de leche",
        plural_text="{amount} ml de leche",
    ),
    "AGUACATE": ExampleFood(
        code="AGUACATE",
        description="aguacate",
        group_code="G",
        amount_per_serving=1,
        use_decimal=False,
        singular_text="1 lonja de aguacate",
        plural_text="{amount} lonjas de aguacate",
    ),
    "ACEITE DE OLIVA": ExampleFood(
        code="ACEITE DE OLIVA",
        description="aceite de oliva",
        group_code="G",
        amount_per_serving=1,
        use_decimal=False,
        singular_text="1 cdta de aceite de oliva",
        plural_text="{amount} cdtas de aceite de oliva",
    ),
    "CAMBUR": ExampleFood(
        code="CAMBUR",
        description="cambur",
        group_code="F",
        amount_per_serving=1,
        use_decimal=False,
        singular_text="1 cambur",
        plural_text="{amount} cambures",
    ),
    "MANZANA": ExampleFood(
        code="MANZANA",
        description="manzana",
        group_code="F",
        amount_per_serving=1,
        use_decimal=False,
        singular_text="1 manzana",
        plural_text="{amount} manzanas",
    ),
    "PERA": ExampleFood(
        code="PERA",
        description="pera",
        group_code="F",
        amount_per_serving=1,
        use_decimal=False,
        singular_text="1 pera",
        plural_text="{amount} peras",
    ),
    "FRESAS": ExampleFood(
        code="FRESAS",
        description="fresas",
        group_code="F",
        amount_per_serving=80,
        use_decimal=False,
        singular_text="80 g de fresas",
        plural_text="{amount} g de fresas",
    ),
    "ENSALADA CRUDA": ExampleFood(
        code="ENSALADA CRUDA",
        description="ensalada cruda",
        group_code="V",
        amount_per_serving=1,
        use_decimal=False,
        singular_text="1 taza de ensalada cruda",
        plural_text="{amount} tazas de ensalada cruda",
    ),
    "VEGETALES SALTEADOS": ExampleFood(
        code="VEGETALES SALTEADOS",
        description="vegetales salteados",
        group_code="V",
        amount_per_serving=1,
        use_decimal=False,
        singular_text="1 taza de vegetales salteados",
        plural_text="{amount} tazas de vegetales salteados",
    ),
}

EXAMPLE_FOOD_ALIASES = {
    "PAN BLANCO": "PAN",
    "PAN INTEGRAL": "PAN",
    "QUESO": "QUESO BLANCO",
    "YOGUR GRIEGO": "YOGURT GRIEGO",
    "ACEITE": "ACEITE DE OLIVA",
    "ENSALADA": "ENSALADA CRUDA",
}


MONTH_NAMES_ES = {
    1: "enero",
    2: "febrero",
    3: "marzo",
    4: "abril",
    5: "mayo",
    6: "junio",
    7: "julio",
    8: "agosto",
    9: "septiembre",
    10: "octubre",
    11: "noviembre",
    12: "diciembre",
}


def require_sheet(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        raise ValidationError(f"No existe la hoja requerida: {sheet_name}")
    return wb[sheet_name]


def to_number(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(",", "."))
        except ValueError:
            return 0.0
    return 0.0


def format_quantity(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def _coerce_excel_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def resolve_patient_age(ws) -> str:
    evaluation_date = _coerce_excel_date(ws["B2"].value)
    birth_date = _coerce_excel_date(ws["C6"].value)
    if evaluation_date is not None and birth_date is not None:
        years = evaluation_date.year - birth_date.year
        if (evaluation_date.month, evaluation_date.day) < (
            birth_date.month,
            birth_date.day,
        ):
            years -= 1
        if years >= 0:
            return str(years)

    return to_age_text(ws["C7"].value)


def load_patient_info(wb) -> PatientInfo:
    ws = require_sheet(wb, "HISTORIA")
    name = str(ws["C4"].value or "").strip()
    ci = str(ws["C5"].value or "").strip()
    sex = str(ws["C10"].value or "").strip()
    discipline = str(ws["I8"].value or "").strip()
    age = resolve_patient_age(ws)

    missing = []
    if not name:
        missing.append("Nombre y Apellido (HISTORIA!C4)")
    if not ci:
        missing.append("Cedula (HISTORIA!C5)")
    if not sex:
        missing.append("Sexo (HISTORIA!C10)")
    if not age:
        missing.append("Edad (HISTORIA!C7)")

    if missing:
        raise ValidationError("Faltan campos: " + "; ".join(missing))

    return PatientInfo(
        name=name,
        ci=ci,
        sex=sex,
        age=age,
        discipline=discipline,
    )


def value_is_missing(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def to_age_text(value) -> str:
    if isinstance(value, (int, float)):
        return str(int(value))
    if value is None:
        return ""
    text = str(value).strip()
    return text


def format_decimal(value, decimals: int = 2, decimal_comma: bool = True) -> str:
    if isinstance(value, (int, float)):
        formatted = f"{float(value):.{decimals}f}"
    else:
        text = str(value or "").strip()
        if not text:
            return ""
        try:
            parsed = float(text.replace(",", "."))
        except ValueError:
            return text
        formatted = f"{parsed:.{decimals}f}"
    if decimal_comma:
        return formatted.replace(".", ",")
    return formatted


def format_table_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return f"{value.day}/{value.month}/{value.year % 100:02d}"
    if isinstance(value, date):
        return f"{value.day}/{value.month}/{value.year % 100:02d}"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value).strip()


def normalize_lookup_label(value: str) -> str:
    normalized = value.replace("_", " ").strip().upper()
    for source, target in (
        ("Á", "A"),
        ("É", "E"),
        ("Í", "I"),
        ("Ó", "O"),
        ("Ú", "U"),
        ("Ü", "U"),
    ):
        normalized = normalized.replace(source, target)
    return " ".join(normalized.split())


@dataclass
class WorkbookFormulaResolver:
    value_wb: object
    formula_wb: object
    cache: Dict[tuple[str, str], object]

    def __init__(self, value_wb, formula_wb) -> None:
        self.value_wb = value_wb
        self.formula_wb = formula_wb
        self.cache = {}

    def resolve_coordinate(
        self,
        sheet_name: str,
        coord: str,
        *,
        visited: set[tuple[str, str]] | None = None,
    ) -> object:
        normalized_coord = coord.replace("$", "")
        cache_key = (sheet_name, normalized_coord)
        if cache_key in self.cache:
            return self.cache[cache_key]
        if visited is None:
            visited = set()
        if cache_key in visited:
            return None
        visited = set(visited)
        visited.add(cache_key)

        value_ws = self.value_wb[sheet_name] if sheet_name in self.value_wb.sheetnames else None
        raw_ws = self.formula_wb[sheet_name] if sheet_name in self.formula_wb.sheetnames else None
        if value_ws is not None:
            cached_value = value_ws[normalized_coord].value
            if not value_is_missing(cached_value):
                self.cache[cache_key] = cached_value
                return cached_value
        if raw_ws is None:
            self.cache[cache_key] = None
            return None

        raw_value = raw_ws[normalized_coord].value
        if value_is_missing(raw_value):
            self.cache[cache_key] = None
            return None
        if not (isinstance(raw_value, str) and raw_value.startswith("=")):
            self.cache[cache_key] = raw_value
            return raw_value

        resolved_value = self._evaluate_formula(
            raw_value[1:],
            current_sheet_name=sheet_name,
            visited=visited,
        )
        self.cache[cache_key] = resolved_value
        return resolved_value

    def _evaluate_formula(
        self,
        expression: str,
        *,
        current_sheet_name: str,
        visited: set[tuple[str, str]],
    ) -> object:
        exact_match = FORMULA_EXACT_REFERENCE_PATTERN.fullmatch(expression.strip())
        if exact_match:
            target_sheet = self._normalize_formula_sheet_name(
                exact_match.group("sheet"),
                default_sheet=current_sheet_name,
            )
            return self.resolve_coordinate(
                target_sheet,
                exact_match.group("coord"),
                visited=visited,
            )

        def replace_reference(match: re.Match[str]) -> str:
            target_sheet = self._normalize_formula_sheet_name(
                match.group("sheet"),
                default_sheet=current_sheet_name,
            )
            resolved = self.resolve_coordinate(
                target_sheet,
                match.group("coord"),
                visited=visited,
            )
            parsed = parse_optional_number(resolved)
            if parsed is None:
                raise ValueError("Referencia no numerica")
            return repr(parsed)

        try:
            python_expression = FORMULA_CELL_REFERENCE_PATTERN.sub(
                replace_reference,
                expression,
            ).replace("^", "**")
            node = ast.parse(python_expression, mode="eval")
            if not self._is_safe_numeric_ast(node):
                return None
            return eval(compile(node, "<formula>", "eval"), {"__builtins__": {}}, {})
        except Exception:
            return None

    def _normalize_formula_sheet_name(
        self,
        raw_sheet_name: str | None,
        *,
        default_sheet: str,
    ) -> str:
        if raw_sheet_name is None:
            return default_sheet
        if raw_sheet_name.startswith("'") and raw_sheet_name.endswith("'"):
            return raw_sheet_name[1:-1].replace("''", "'")
        return raw_sheet_name

    def _is_safe_numeric_ast(self, node: ast.AST) -> bool:
        if isinstance(node, ast.Expression):
            return self._is_safe_numeric_ast(node.body)
        if isinstance(node, ast.Constant):
            return isinstance(node.value, (int, float))
        if isinstance(node, ast.UnaryOp):
            return isinstance(node.op, (ast.UAdd, ast.USub)) and self._is_safe_numeric_ast(node.operand)
        if isinstance(node, ast.BinOp):
            return (
                isinstance(node.op, (ast.Add, ast.Sub, ast.Mult, ast.Div, ast.Pow))
                and self._is_safe_numeric_ast(node.left)
                and self._is_safe_numeric_ast(node.right)
            )
        return False


def hydrate_formula_cells_for_inspection(value_wb, formula_wb) -> None:
    resolver = WorkbookFormulaResolver(value_wb, formula_wb)
    for sheet_name in INSPECTION_SHEETS:
        if sheet_name not in formula_wb.sheetnames or sheet_name not in value_wb.sheetnames:
            continue
        formula_ws = formula_wb[sheet_name]
        value_ws = value_wb[sheet_name]
        for row in formula_ws.iter_rows():
            for formula_cell in row:
                raw_value = formula_cell.value
                if not (isinstance(raw_value, str) and raw_value.startswith("=")):
                    continue
                value_cell = value_ws[formula_cell.coordinate]
                if not value_is_missing(value_cell.value):
                    continue
                resolved_value = resolver.resolve_coordinate(
                    sheet_name,
                    formula_cell.coordinate,
                )
                if not value_is_missing(resolved_value):
                    value_cell.value = resolved_value


def load_workbook_for_inspection(excel_path: Path | str):
    excel_path = Path(excel_path)
    keep_vba = excel_path.suffix.lower() == ".xlsm"
    value_wb = load_workbook(excel_path, data_only=True, keep_vba=keep_vba)
    formula_wb = load_workbook(excel_path, data_only=False, keep_vba=keep_vba)
    try:
        hydrate_formula_cells_for_inspection(value_wb, formula_wb)
    finally:
        formula_wb.close()
    return value_wb


def build_sheet_headers(ws, header_row: int = 1) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        headers[normalize_lookup_label(str(cell.value))] = cell.column
    return headers


def empty_meal_distribution() -> Dict[str, Dict[str, float]]:
    return {
        meal_def["name"]: {group_code: 0.0 for group_code in GROUP_ROWS}
        for meal_def in MEAL_DEFS
    }


def normalize_food_name(value: str) -> str:
    normalized = normalize_lookup_label(value)
    return EXAMPLE_FOOD_ALIASES.get(normalized, normalized)


def build_food_lookup_keys(value: str) -> List[str]:
    exact_key = normalize_lookup_label(value)
    alias_key = EXAMPLE_FOOD_ALIASES.get(exact_key)
    if alias_key and alias_key != exact_key:
        return [exact_key, alias_key]
    return [exact_key]


def parse_bool_like(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    normalized = normalize_lookup_label(str(value))
    return normalized in {"SI", "S", "YES", "Y", "TRUE", "1"}


def parse_float_like(value, field_label: str) -> float:
    if value_is_missing(value):
        raise ValidationError(f"Falta campo: {field_label}")
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    try:
        return float(text)
    except ValueError as exc:
        raise ValidationError(
            f"Valor inválido en {field_label}: {value}"
        ) from exc


def format_example_amount(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.2f}".rstrip("0").rstrip(".").replace(".", ",")


def build_example_fragment(food: ExampleFood, servings: float) -> str:
    amount = servings * food.amount_per_serving
    if amount == food.amount_per_serving:
        return food.singular_text
    rendered_amount = format_example_amount(amount)
    return food.plural_text.replace("{amount}", rendered_amount).replace("{n}", rendered_amount)


def build_example_food_lookup(wb) -> Dict[str, ExampleFood]:
    if "EQUIVALENCIAS_EJEMPLOS" not in wb.sheetnames:
        return DEFAULT_EXAMPLE_FOOD_DEFS.copy()

    ws = wb["EQUIVALENCIAS_EJEMPLOS"]
    headers: Dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        headers[normalize_lookup_label(str(cell.value))] = cell.column

    required_headers = [
        "CODIGO ALIMENTO",
        "GRUPO",
        "CANTIDAD POR RACION",
        "TEXTO SINGULAR",
        "TEXTO PLURAL",
    ]
    missing_headers = [
        header for header in required_headers if header not in headers]
    if missing_headers:
        raise ValidationError(
            "Faltan columnas en EQUIVALENCIAS_EJEMPLOS: " +
            ", ".join(missing_headers)
        )

    food_lookup: Dict[str, ExampleFood] = {}
    duplicated_keys: set[str] = set()

    for row_idx in range(2, ws.max_row + 1):
        code_value = ws.cell(
            row=row_idx, column=headers["CODIGO ALIMENTO"]).value
        if value_is_missing(code_value):
            continue

        code = normalize_lookup_label(str(code_value))
        group_value = ws.cell(row=row_idx, column=headers["GRUPO"]).value
        description_col = headers.get("DESCRIPCION BASE")
        description_value = ws.cell(
            row=row_idx, column=description_col).value if description_col else code_value
        quantity_value = ws.cell(
            row=row_idx, column=headers["CANTIDAD POR RACION"]).value
        singular_value = ws.cell(
            row=row_idx, column=headers["TEXTO SINGULAR"]).value
        plural_value = ws.cell(
            row=row_idx, column=headers["TEXTO PLURAL"]).value
        decimal_col = headers.get("USA DECIMAL")
        decimal_value = ws.cell(
            row=row_idx, column=decimal_col).value if decimal_col else None

        if value_is_missing(group_value):
            raise ValidationError(
                f"Falta campo: GRUPO (EQUIVALENCIAS_EJEMPLOS!B{row_idx})"
            )
        if value_is_missing(singular_value):
            raise ValidationError(
                f"Falta campo: TEXTO_SINGULAR (EQUIVALENCIAS_EJEMPLOS!G{row_idx})"
            )
        if value_is_missing(plural_value):
            raise ValidationError(
                f"Falta campo: TEXTO_PLURAL (EQUIVALENCIAS_EJEMPLOS!H{row_idx})"
            )

        group_label = normalize_lookup_label(str(group_value))
        if group_label not in EXAMPLE_GROUP_HEADERS:
            raise ValidationError(
                f"Grupo no soportado en EQUIVALENCIAS_EJEMPLOS fila {row_idx}: {group_value}"
            )

        food = ExampleFood(
            code=code,
            description=str(description_value or code_value).strip(),
            group_code=EXAMPLE_GROUP_HEADERS[group_label],
            amount_per_serving=parse_float_like(
                quantity_value,
                f"CANTIDAD_POR_RACION (EQUIVALENCIAS_EJEMPLOS!D{row_idx})",
            ),
            use_decimal=parse_bool_like(decimal_value),
            singular_text=str(singular_value).strip(),
            plural_text=str(plural_value).strip(),
        )

        lookup_keys = {
            code,
            normalize_lookup_label(str(description_value or code_value)),
        }
        for lookup_key in lookup_keys:
            if not lookup_key:
                continue
            if lookup_key in duplicated_keys:
                continue
            if lookup_key in food_lookup:
                duplicated_keys.add(lookup_key)
                food_lookup.pop(lookup_key, None)
                continue
            food_lookup[lookup_key] = food

    if not food_lookup:
        return DEFAULT_EXAMPLE_FOOD_DEFS.copy()

    return food_lookup


def load_examples_sheet(wb) -> Dict[str, Dict[str, str]]:
    if "EJEMPLOS_COMIDAS" not in wb.sheetnames:
        return {}

    ws = wb["EJEMPLOS_COMIDAS"]
    headers: Dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        headers[normalize_lookup_label(str(cell.value))] = cell.column

    if "COMIDA" not in headers:
        raise ValidationError(
            "La hoja EJEMPLOS_COMIDAS debe incluir una columna COMIDA en la fila 1"
        )

    meal_rows: Dict[str, Dict[str, str]] = {}
    for row_idx in range(2, ws.max_row + 1):
        meal_value = ws.cell(row=row_idx, column=headers["COMIDA"]).value
        if value_is_missing(meal_value):
            continue
        meal_name = str(meal_value).strip().upper()
        if meal_name not in {meal["name"] for meal in MEAL_DEFS}:
            raise ValidationError(
                f"Comida no reconocida en EJEMPLOS_COMIDAS!A{row_idx}: {meal_value}"
            )
        if meal_name in meal_rows:
            raise ValidationError(
                f"La comida {meal_name} está repetida en EJEMPLOS_COMIDAS"
            )

        row_data: Dict[str, str] = {}
        for header_name, group_code in EXAMPLE_GROUP_HEADERS.items():
            col_idx = headers.get(header_name)
            value = ws.cell(
                row=row_idx, column=col_idx).value if col_idx else None
            row_data[group_code] = str(value).strip() if value else ""

        obs_col = headers.get("OBSERVACION")
        obs_value = ws.cell(
            row=row_idx, column=obs_col).value if obs_col else None
        row_data["OBSERVACION"] = str(obs_value).strip() if obs_value else ""
        meal_rows[meal_name] = row_data

    return meal_rows


def is_guide_example_row(row_data: Dict[str, str]) -> bool:
    observation = normalize_lookup_label(row_data.get("OBSERVACION", ""))
    if observation == normalize_lookup_label(EXAMPLE_GUIDE_OBSERVATION):
        return True

    meaningful_values = [
        value.strip().lower()
        for key, value in row_data.items()
        if key != "OBSERVACION" and value.strip()
    ]
    return bool(meaningful_values) and all(
        value in EXAMPLE_GUIDE_VALUES
        for value in meaningful_values
    )


def load_plan_template_distribution(wb) -> Dict[str, Dict[str, float]]:
    ws = require_sheet(wb, PLAN_TEMPLATE_SHEET)
    headers = build_sheet_headers(ws)
    required_headers = ["COMIDA", *EXAMPLE_GROUP_HEADERS.keys()]
    missing_headers = [
        header for header in required_headers if header not in headers]
    if missing_headers:
        raise ValidationError(
            f"Faltan columnas en {PLAN_TEMPLATE_SHEET}: " +
            ", ".join(missing_headers)
        )

    distribution = empty_meal_distribution()
    seen_meals: set[str] = set()

    for row_idx in range(2, ws.max_row + 1):
        meal_value = ws.cell(row=row_idx, column=headers["COMIDA"]).value
        if value_is_missing(meal_value):
            continue

        meal_name = normalize_lookup_label(str(meal_value))
        if meal_name not in distribution:
            raise ValidationError(
                f"Comida no reconocida en {PLAN_TEMPLATE_SHEET}!A{row_idx}: {meal_value}"
            )
        if meal_name in seen_meals:
            raise ValidationError(
                f"La comida {meal_name} está repetida en {PLAN_TEMPLATE_SHEET}"
            )

        for header_name, group_code in EXAMPLE_GROUP_HEADERS.items():
            col_idx = headers[header_name]
            distribution[meal_name][group_code] = to_number(
                ws.cell(row=row_idx, column=col_idx).value
            )

        seen_meals.add(meal_name)

    return distribution


def load_meal_distribution(wb) -> Dict[str, Dict[str, float]]:
    return load_plan_template_distribution(wb)


def build_label_lookup(rows: List[Tuple[str, object]]) -> Dict[str, object]:
    return {
        normalize_lookup_label(label): value
        for label, value in rows
        if label.strip()
    }


def value_from_lookup(
    lookup: Dict[str, object],
    labels: List[str],
) -> object | None:
    for label in labels:
        normalized = normalize_lookup_label(label)
        if normalized not in lookup:
            continue
        candidate = lookup[normalized]
        if isinstance(candidate, (list, tuple)):
            for item in reversed(candidate):
                if not value_is_missing(item):
                    return item
            continue
        if not value_is_missing(candidate):
            return candidate
    return None


def lookup_contains_any_label(
    lookup: Dict[str, object],
    labels: List[str],
) -> bool:
    return any(normalize_lookup_label(label) in lookup for label in labels)


ANTHRO_PESO_LABELS = [
    "Peso (Kg)",
    "Peso actual (kg)",
    "Peso Actual (Kg)",
]

ANTHRO_TALLA_M_LABELS = ["Talla (m)"]

ANTHRO_MASA_MAGRA_LABELS = ["Kg de Masa Magra"]

ANTHRO_MASA_GRASA_LABELS = ["Kg de Grasa"]

ANTHRO_PCT_GRASA_CARTER_LABELS = [
    "% Grasa (Carter 1986)",
    "% Grasa Carter",
    "% Grasa Carter 1986",
    "%grasa carter",
]

ANTHRO_REQUIRED_VALUE_SUMMARY_FIELDS = [
    ("Fecha", ["Fecha"], "date"),
    ("Peso (Kg)", ANTHRO_PESO_LABELS, "positive_number"),
    ("Talla parada (cm)", ["Talla parada (cm)"], "positive_number"),
    ("% Grasa (Carter 1986)", ANTHRO_PCT_GRASA_CARTER_LABELS, "positive_number"),
    ("Kg de Masa Magra", ANTHRO_MASA_MAGRA_LABELS, "positive_number"),
    ("Kg de Grasa", ANTHRO_MASA_GRASA_LABELS, "positive_number"),
]

ANTHRO_REQUIRED_VALUE_MEASUREMENT_FIELDS = [
    ("Fecha de evaluación", ["Fecha de evaluación"], "date"),
    ("Talla (m)", ANTHRO_TALLA_M_LABELS, "positive_number"),
]


def detect_anthro_value_columns(ws) -> List[int]:
    last_value_col = 0
    for col_idx in range(3, ws.max_column + 1):
        if any(
            not value_is_missing(ws.cell(row=row_idx, column=col_idx).value)
            for row_idx in range(1, ws.max_row + 1)
        ):
            last_value_col = col_idx
    if last_value_col == 0:
        return []
    return list(range(3, last_value_col + 1))


def parse_optional_number(value) -> float | None:
    if value_is_missing(value) or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def anthro_required_value_is_valid(value, expected_type: str) -> bool:
    if expected_type == "date":
        if isinstance(value, (datetime, date)):
            return True
        return isinstance(value, str) and bool(value.strip())
    if expected_type == "positive_number":
        parsed = parse_optional_number(value)
        return parsed is not None and parsed > 0
    return not value_is_missing(value)


def value_list_from_lookup(
    lookup: Dict[str, object],
    labels: List[str],
) -> List[object]:
    for label in labels:
        normalized = normalize_lookup_label(label)
        if normalized not in lookup:
            continue
        candidate = lookup[normalized]
        if isinstance(candidate, (list, tuple)):
            return list(candidate)
        return [candidate]
    return []


def filter_anthro_rows_by_indices(
    rows: List[Tuple[str, List[object]]],
    indices: List[int],
) -> List[Tuple[str, List[object]]]:
    return [
        (
            label,
            [values[idx] if idx < len(values) else None for idx in indices],
        )
        for label, values in rows
    ]


def find_valid_anthro_column_indices(
    summary_rows: List[Tuple[str, List[object]]],
    measurement_rows: List[Tuple[str, List[object]]],
) -> List[int]:
    max_len = max(
        [len(values) for _, values in summary_rows + measurement_rows],
        default=0,
    )
    if max_len == 0:
        return []

    summary_lookup = build_label_lookup_first(summary_rows)
    measurement_lookup = build_label_lookup_first(measurement_rows)
    valid_indices: List[int] = []

    for idx in range(max_len):
        summary_ok = all(
            idx < len(value_list_from_lookup(summary_lookup, labels))
            and anthro_required_value_is_valid(
                value_list_from_lookup(summary_lookup, labels)[idx],
                expected_type,
            )
            for _, labels, expected_type in ANTHRO_REQUIRED_VALUE_SUMMARY_FIELDS
        )
        measurement_ok = all(
            idx < len(value_list_from_lookup(measurement_lookup, labels))
            and anthro_required_value_is_valid(
                value_list_from_lookup(measurement_lookup, labels)[idx],
                expected_type,
            )
            for _, labels, expected_type in ANTHRO_REQUIRED_VALUE_MEASUREMENT_FIELDS
        )
        if summary_ok and measurement_ok:
            valid_indices.append(idx)

    return valid_indices


def sanitize_anthro_rows(
    summary_rows: List[Tuple[str, List[object]]],
    measurement_rows: List[Tuple[str, List[object]]],
) -> Tuple[List[Tuple[str, List[object]]], List[Tuple[str, List[object]]]]:
    summary_rows = normalize_anthro_rows(summary_rows)
    measurement_rows = normalize_anthro_rows(measurement_rows)
    valid_indices = find_valid_anthro_column_indices(
        summary_rows,
        measurement_rows,
    )
    if not valid_indices:
        return summary_rows, measurement_rows
    return (
        filter_anthro_rows_by_indices(summary_rows, valid_indices),
        filter_anthro_rows_by_indices(measurement_rows, valid_indices),
    )


def normalize_anthro_rows(
    rows: List[Tuple[str, List[object]]],
) -> List[Tuple[str, List[object]]]:
    if not rows:
        return []

    max_len = max(len(values) for _, values in rows)
    if max_len == 0:
        return [(label, []) for label, _ in rows]

    last_used_idx = 0
    for idx in range(max_len):
        if any(
            idx < len(values) and not value_is_missing(values[idx])
            for _, values in rows
        ):
            last_used_idx = idx + 1
    if last_used_idx == 0:
        last_used_idx = max_len

    normalized_rows: List[Tuple[str, List[object]]] = []
    for label, values in rows:
        padded = list(values[:last_used_idx])
        padded.extend([None] * (last_used_idx - len(padded)))
        normalized_rows.append((label, padded))
    return normalized_rows


def build_display_anthro_rows(
    rows: List[Tuple[str, List[object]]],
) -> List[List[str]]:
    return [
        [
            format_table_value(label),
            *[format_table_value(value) for value in values],
        ]
        for label, values in rows
    ]


def anthropometric_data_from_rows(
    patient: PatientInfo,
    summary_rows_raw: List[Tuple[str, List[object]]],
    measurement_rows_raw: List[Tuple[str, List[object]]],
) -> AnthropometricReportData:
    summary_rows_raw, measurement_rows_raw = sanitize_anthro_rows(
        summary_rows_raw,
        measurement_rows_raw,
    )
    summary_lookup = build_label_lookup(summary_rows_raw)
    measurement_lookup = build_label_lookup(measurement_rows_raw)

    peso_corporal_value = value_from_lookup(summary_lookup, ANTHRO_PESO_LABELS)
    if value_is_missing(peso_corporal_value):
        raise ValidationError(
            "Falta campo: Peso corporal en la tabla resumen antropométrica"
        )

    estatura_value = value_from_lookup(
        measurement_lookup, ANTHRO_TALLA_M_LABELS)
    if value_is_missing(estatura_value):
        raise ValidationError(
            "Falta campo: Talla (m) en la tabla de medidas antropométricas"
        )

    masa_magra_value = value_from_lookup(
        summary_lookup, ANTHRO_MASA_MAGRA_LABELS)
    if value_is_missing(masa_magra_value):
        raise ValidationError(
            "Falta campo: Kg de Masa Magra en la tabla resumen antropomÃ©trica"
        )

    masa_grasa_value = value_from_lookup(
        summary_lookup, ANTHRO_MASA_GRASA_LABELS)
    if value_is_missing(masa_grasa_value):
        raise ValidationError(
            "Falta campo: Kg de Grasa en la tabla resumen antropométrica"
        )

    pct_grasa_value = value_from_lookup(
        summary_lookup, ANTHRO_PCT_GRASA_CARTER_LABELS)
    if value_is_missing(pct_grasa_value):
        raise ValidationError(
            "Falta campo: % Grasa (Carter 1986) en la tabla resumen antropométrica"
        )

    return AnthropometricReportData(
        patient=patient,
        peso_corporal_kg=format_decimal(peso_corporal_value),
        estatura_m=format_decimal(estatura_value),
        masa_magra_kg=format_decimal(masa_magra_value),
        masa_grasa_kg=format_decimal(masa_grasa_value),
        pct_grasa_carter=format_decimal(pct_grasa_value),
        table_resumen=build_display_anthro_rows(summary_rows_raw),
        table_medidas=build_display_anthro_rows(measurement_rows_raw),
    )


def load_anthropometric_template_data(
    wb,
    patient: PatientInfo,
) -> AnthropometricReportData:
    ws = require_sheet(wb, ANTHRO_TEMPLATE_SHEET)
    headers = build_sheet_headers(ws)
    required_headers = ["SECCION", "ETIQUETA"]
    missing_headers = [
        header for header in required_headers if header not in headers]
    if missing_headers:
        raise ValidationError(
            f"Faltan columnas en {ANTHRO_TEMPLATE_SHEET}: " +
            ", ".join(missing_headers)
        )
    value_columns = detect_anthro_value_columns(ws)
    if not value_columns:
        raise ValidationError(
            f"La hoja {ANTHRO_TEMPLATE_SHEET} debe incluir al menos una columna de valores desde la columna C"
        )

    summary_rows_raw: List[Tuple[str, List[object]]] = []
    measurement_rows_raw: List[Tuple[str, List[object]]] = []

    for row_idx in range(2, ws.max_row + 1):
        section_value = ws.cell(row=row_idx, column=headers["SECCION"]).value
        label_value = ws.cell(row=row_idx, column=headers["ETIQUETA"]).value

        if value_is_missing(section_value) and value_is_missing(label_value):
            continue
        if value_is_missing(section_value) or value_is_missing(label_value):
            raise ValidationError(
                f"Cada fila de {ANTHRO_TEMPLATE_SHEET} requiere SECCION y ETIQUETA"
            )

        section = normalize_lookup_label(str(section_value))
        row_values = [
            ws.cell(row=row_idx, column=col_idx).value
            for col_idx in value_columns
        ]
        row = (str(label_value).strip(), row_values)
        if section == "RESUMEN":
            summary_rows_raw.append(row)
        elif section in {"MEDIDAS", "MEDIDA"}:
            measurement_rows_raw.append(row)
        else:
            raise ValidationError(
                f"Sección no reconocida en {ANTHRO_TEMPLATE_SHEET}!A{row_idx}: {section_value}"
            )

    if not summary_rows_raw or not measurement_rows_raw:
        raise ValidationError(
            f"La hoja {ANTHRO_TEMPLATE_SHEET} debe incluir filas de RESUMEN y MEDIDAS"
        )

    summary_rows_raw = normalize_anthro_rows(summary_rows_raw)
    measurement_rows_raw = normalize_anthro_rows(measurement_rows_raw)

    return anthropometric_data_from_rows(
        patient=patient,
        summary_rows_raw=summary_rows_raw,
        measurement_rows_raw=measurement_rows_raw,
    )


def load_anthropometric_data(wb) -> AnthropometricReportData:
    patient = load_patient_info(wb)
    return load_anthropometric_template_data(wb, patient)


def month_name_es(reference_date: date) -> str:
    return MONTH_NAMES_ES[reference_date.month]


def build_anthropometric_replacements(
    data: AnthropometricReportData, today: date
) -> Dict[str, str]:
    next_control = today + timedelta(days=42)
    discipline = data.patient.discipline or "____________________"
    return {
        "{{PACIENTE}}": format_short_name(data.patient.name),
        "{{EDAD}}": data.patient.age,
        "{{CI}}": data.patient.ci,
        "{{DISCIPLINA}}": discipline,
        "{{OBJETIVO}}": "PERDER GRASA",
        "{{PESO_CORPORAL_KG}}": data.peso_corporal_kg,
        "{{ESTATURA_M}}": data.estatura_m,
        "{{MASA_MAGRA_KG}}": data.masa_magra_kg,
        "{{MASA_GRASA_KG}}": data.masa_grasa_kg,
        "{{PCT_GRASA_CARTER}}": data.pct_grasa_carter,
        "{{MES_ACTUAL}}": month_name_es(today),
        "{{PROXIMO_CONTROL}}": next_control.strftime("%d/%m/%Y"),
    }


def build_summary_table_replacements(data: AnthropometricReportData) -> Dict[str, str]:
    replacements: Dict[str, str] = {}
    for row_idx, row_values in enumerate(data.table_resumen, start=1):
        for col_idx, cell_value in enumerate(row_values, start=1):
            replacements[f"{{{{R{row_idx}C{col_idx}}}}}"] = cell_value
    return replacements


def build_measurements_table_replacements(
    data: AnthropometricReportData,
) -> Dict[str, str]:
    replacements: Dict[str, str] = {}
    for row_idx, row_values in enumerate(data.table_medidas, start=1):
        for col_idx, cell_value in enumerate(row_values, start=1):
            replacements[f"{{{{M{row_idx}C{col_idx}}}}}"] = cell_value
            replacements[f"{{{{R{row_idx}C{col_idx}}}}}"] = cell_value
    return replacements


def format_short_name(full_name: str) -> str:
    parts = [p for p in full_name.split() if p.strip()]
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0]
    if len(parts) >= 3:
        return f"{parts[0]} {parts[-2]}"
    return f"{parts[0]} {parts[1]}"


def build_replacements(patient: PatientInfo) -> Dict[str, str]:
    placeholder = "____________________"
    display_name = format_short_name(patient.name)
    return {
        "{{PACIENTE}}": display_name,
        "{{DISCIPLINA}}": patient.discipline or placeholder,
        "{{OBJETIVO}}": placeholder,
        "{{SEXO}}": patient.sex,
        "{{EDAD}}": patient.age,
    }


def build_meal_replacements(
    meal_distribution: Dict[str, Dict[str, float]], meal_def
) -> Tuple[Dict[str, str], Dict[str, float], bool, List[str], Dict[str, float]]:
    values = {
        code: to_number(meal_distribution.get(meal_def["name"], {}).get(code))
        for code in GROUP_ROWS
    }

    replacements = {}
    for code, suffix in GROUP_SUFFIX.items():
        placeholder = f"{{{{{meal_def['name']}_{suffix}}}}}"
        replacements[placeholder] = "" if values[code] == 0 else format_quantity(
            values[code]
        )

    include = any(values[code] > 0 for code in meal_def["groups"])
    tokens = [
        f"{{{{{meal_def['name']}_{GROUP_SUFFIX[code]}}}}}"
        for code in meal_def["groups"]
    ]
    placeholder_values = {
        f"{{{{{meal_def['name']}_{suffix}}}}}": values[code]
        for code, suffix in GROUP_SUFFIX.items()
    }
    return replacements, values, include, tokens, placeholder_values


def build_meal_example_texts(
    wb, meal_distribution: Dict[str, Dict[str, float]]
) -> Dict[str, str]:
    meal_rows = load_examples_sheet(wb)
    if not meal_rows:
        return {}
    food_lookup = build_example_food_lookup(wb)

    example_texts: Dict[str, str] = {}
    for meal_def in MEAL_DEFS:
        meal_name = meal_def["name"]
        servings = meal_distribution.get(
            meal_name,
            {group_code: 0.0 for group_code in GROUP_ROWS},
        )
        needs_example = any(
            servings[group_code] > 0 for group_code in meal_def["groups"]
        )
        if meal_name not in meal_rows:
            continue

        row_data = meal_rows[meal_name]
        if is_guide_example_row(row_data):
            continue

        fragments: List[str] = []
        for group_code in MEAL_EXAMPLE_ORDER.get(meal_name, meal_def["groups"]):
            if group_code not in meal_def["groups"]:
                continue
            serving_count = servings[group_code]
            if serving_count <= 0:
                continue

            food_name = row_data.get(group_code, "")
            if not food_name:
                continue

            food = None
            for lookup_key in build_food_lookup_keys(food_name):
                food = food_lookup.get(lookup_key)
                if food is not None:
                    break
            if food is None:
                fragments.append(food_name)
            else:
                fragments.append(build_example_fragment(food, serving_count))

        observation = row_data.get("OBSERVACION", "")
        if not fragments and not observation and not needs_example:
            continue

        if fragments:
            example_text = "EJEMPLO: " + " + ".join(fragments)
            if observation:
                example_text += f" | {observation}"
        elif needs_example and observation:
            example_text = f"NOTA: {observation}"
        else:
            continue
        example_texts[meal_name] = example_text

    return example_texts


def build_totals_replacements(
    meal_distribution: Dict[str, Dict[str, float]]
) -> Dict[str, str]:
    totals = {
        group_code: sum(
            meal_distribution[meal_name][group_code]
            for meal_name in meal_distribution
        )
        for group_code in GROUP_ROWS
    }
    return {
        "{{TOTAL_LACTEOS}}": "" if totals["L"] == 0 else format_quantity(totals["L"]),
        "{{TOTAL_VEGETALES}}": "" if totals["V"] == 0 else format_quantity(totals["V"]),
        "{{TOTAL_FRUTAS}}": "" if totals["F"] == 0 else format_quantity(totals["F"]),
        "{{TOTAL_ALMIDONES}}": "" if totals["A"] == 0 else format_quantity(totals["A"]),
        "{{TOTAL_PROTEINAS}}": "" if totals["P"] == 0 else format_quantity(totals["P"]),
        "{{TOTAL_GRASAS}}": "" if totals["G"] == 0 else format_quantity(totals["G"]),
    }


@dataclass
class ValidationIssue:
    section: str
    message: str
    sheet: str = ""
    location: str = ""
    field: str = ""
    expected: str = ""
    actual_value: str = ""
    severity: str = "error"

    @property
    def is_blocking(self) -> bool:
        return self.severity == "error"


@dataclass
class ParsedWorkbookData:
    patient: PatientInfo
    meal_distribution: Dict[str, Dict[str, float]]
    meal_totals: Dict[str, float]
    anthro_data: AnthropometricReportData
    meal_examples: Dict[str, str]
    issues: List[ValidationIssue]
    examples_status: str

    @property
    def has_blocking_issues(self) -> bool:
        return any(issue.is_blocking for issue in self.issues)


SECTION_PATIENT = "Paciente / HISTORIA"
SECTION_PLAN = "Plan / PLAN_ALIMENTACION_TEMPLATE"
SECTION_ANTHRO = "Antropometria / ANTROPOMETRIA_TEMPLATE"
SECTION_EXAMPLES = "Ejemplos / EJEMPLOS_COMIDAS"
SECTION_ORDER = [
    SECTION_PATIENT,
    SECTION_PLAN,
    SECTION_ANTHRO,
    SECTION_EXAMPLES,
]

ANTHRO_REQUIRED_SUMMARY_FIELDS = [
    ("Evaluación", ["Evaluación"]),
    ("Fecha", ["Fecha"]),
    ("Peso (Kg)", ANTHRO_PESO_LABELS),
    ("Talla parada (cm)", ["Talla parada (cm)"]),
    ("% Grasa (Carter 1986)", ANTHRO_PCT_GRASA_CARTER_LABELS),
    ("% Grasa (Durnin y W. 1974)", ["% Grasa (Durnin y W. 1974)"]),
    ("Interpretación", ["Interpretación"]),
    ("Kg de Masa Magra", ["Kg de Masa Magra"]),
    ("Kg de Grasa", ANTHRO_MASA_GRASA_LABELS),
    ("Masa Muscular (Kg)", ["Masa Muscular (Kg)"]),
    ("Masa Adiposa (Kg)", ["Masa Adiposa (Kg)"]),
    ("Sumatoria de 6 pliegues", ["Sumatoria de 6 pliegues"]),
    ("Somatotipo", ["Somatotipo"]),
]

ANTHRO_REQUIRED_MEASUREMENT_FIELDS = [
    ("Fecha de evaluación", ["Fecha de evaluación"]),
    ("Peso actual (kg)", ["Peso actual (kg)"]),
    ("Talla (m)", ANTHRO_TALLA_M_LABELS),
    ("Talla (cm)", ["Talla (cm)"]),
    ("Circunferencias(cm)", ["Circunferencias(cm)"]),
    ("Brazo relajado (cm)", ["Brazo relajado (cm)"]),
    ("Brazo Flexionado en Tensión (cm)", ["Brazo Flexionado en Tensión (cm)"]),
    ("Antebrazo máximo (cm)", ["Antebrazo máximo (cm)"]),
    ("Tórax (Mesoesternal) (cm)", ["Tórax (Mesoesternal) (cm)"]),
    ("Cintura mínimo (cm)", ["Cintura mínimo (cm)"]),
    ("Cadera máximo (cm)", ["Cadera máximo (cm)"]),
    ("Muslo máximo (cm)", ["Muslo máximo (cm)"]),
    ("Muslo medial (cm)", ["Muslo medial (cm)"]),
    ("Pantorrilla máximo (cm)", ["Pantorrilla máximo (cm)"]),
    ("Pliegues (mm)", ["Pliegues (mm)"]),
    ("Bíceps (mm)", ["Bíceps (mm)"]),
    ("Tríceps (mm)", ["Tríceps (mm)"]),
    ("Subescapular (mm)", ["Subescapular (mm)"]),
    ("Ileo-crestal (mm)", ["Ileo-crestal (mm)"]),
    ("Supra-espinal (mm)", ["Supra-espinal (mm)"]),
    ("Abdominal (mm)", ["Abdominal (mm)"]),
    ("Muslo frontal (mm)", ["Muslo frontal (mm)"]),
    ("Pantorrilla (mm)", ["Pantorrilla (mm)"]),
    ("Diametros óseos (cm)", ["Diametros óseos (cm)"]),
    ("Humeral (Biepicondilar)", ["Humeral (Biepicondilar)"]),
    ("Femoral (Biepicondilar)", ["Femoral (Biepicondilar)"]),
]


def blank_patient_info() -> PatientInfo:
    return PatientInfo(name="", ci="", sex="", age="", discipline="")


def blank_anthro_data(patient: PatientInfo) -> AnthropometricReportData:
    return AnthropometricReportData(
        patient=patient,
        peso_corporal_kg="",
        estatura_m="",
        masa_magra_kg="",
        masa_grasa_kg="",
        pct_grasa_carter="",
        table_resumen=[],
        table_medidas=[],
    )


def make_issue(
    section: str,
    message: str,
    sheet: str = "",
    location: str = "",
    field: str = "",
    expected: str = "",
    actual_value: str = "",
    severity: str = "error",
) -> ValidationIssue:
    return ValidationIssue(
        section=section,
        message=message,
        sheet=sheet,
        location=location,
        field=field,
        expected=expected,
        actual_value=actual_value,
        severity=severity,
    )


def blocking_issues_for_sections(
    issues: List[ValidationIssue],
    sections: List[str] | tuple[str, ...] | set[str],
) -> List[ValidationIssue]:
    section_names = set(sections)
    return [
        issue
        for issue in issues
        if issue.section in section_names and issue.is_blocking
    ]


def format_actual_value(value) -> str:
    rendered = format_table_value(value)
    return rendered if rendered else "vacio"


def parse_number_with_issue(
    value,
    *,
    section: str,
    sheet: str,
    location: str,
    field: str,
    issues: List[ValidationIssue],
) -> float:
    if value_is_missing(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    try:
        return float(text)
    except ValueError:
        issues.append(
            make_issue(
                section=section,
                message=f"El valor de {field} debe ser numerico.",
                sheet=sheet,
                location=location,
                field=field,
                expected="numero",
                actual_value=format_actual_value(value),
            )
        )
        return 0.0


def calculate_meal_totals(
    meal_distribution: Dict[str, Dict[str, float]]
) -> Dict[str, float]:
    return {
        group_code: sum(
            meal_distribution[meal_name][group_code]
            for meal_name in meal_distribution
        )
        for group_code in GROUP_ROWS
    }


def inspect_patient_info(wb, issues: List[ValidationIssue]) -> PatientInfo:
    if "HISTORIA" not in wb.sheetnames:
        issues.append(
            make_issue(
                section=SECTION_PATIENT,
                message="Falta la hoja HISTORIA.",
                sheet="HISTORIA",
                expected="Hoja HISTORIA con datos del paciente",
            )
        )
        return blank_patient_info()

    ws = wb["HISTORIA"]
    name = str(ws["C4"].value or "").strip()
    ci = str(ws["C5"].value or "").strip()
    sex = str(ws["C10"].value or "").strip()
    discipline = str(ws["I8"].value or "").strip()
    age = resolve_patient_age(ws)

    required_fields = [
        ("Nombre y Apellido", "C4", name),
        ("Cedula", "C5", ci),
        ("Edad", "C7", age),
        ("Sexo", "C10", sex),
    ]
    for field_name, location, value in required_fields:
        if value_is_missing(value):
            severity = "warning" if field_name == "Edad" else "error"
            issues.append(
                make_issue(
                    section=SECTION_PATIENT,
                    message=f"Falta el campo {field_name}.",
                    sheet="HISTORIA",
                    location=location,
                    field=field_name,
                    expected="valor no vacio",
                    actual_value="vacio",
                    severity=severity,
                )
            )

    return PatientInfo(
        name=name,
        ci=ci,
        sex=sex,
        age=age,
        discipline=discipline,
    )


def inspect_plan_distribution(
    wb,
    issues: List[ValidationIssue],
) -> Dict[str, Dict[str, float]]:
    distribution = empty_meal_distribution()
    if PLAN_TEMPLATE_SHEET not in wb.sheetnames:
        issues.append(
            make_issue(
                section=SECTION_PLAN,
                message=f"Falta la hoja obligatoria {PLAN_TEMPLATE_SHEET}.",
                sheet=PLAN_TEMPLATE_SHEET,
                expected="Hoja con columnas COMIDA, LACTEOS, VEGETALES, FRUTAS, ALMIDONES, PROTEINAS, GRASAS",
            )
        )
        return distribution

    ws = wb[PLAN_TEMPLATE_SHEET]
    headers = build_sheet_headers(ws)
    required_headers = ["COMIDA", *EXAMPLE_GROUP_HEADERS.keys()]
    missing_headers = [
        header for header in required_headers if header not in headers]
    if missing_headers:
        issues.append(
            make_issue(
                section=SECTION_PLAN,
                message="Faltan columnas obligatorias en la hoja del plan.",
                sheet=PLAN_TEMPLATE_SHEET,
                field="columnas",
                expected=", ".join(required_headers),
                actual_value=", ".join(
                    sorted(headers.keys())) or "sin encabezados",
            )
        )

    meal_col = headers.get("COMIDA")
    seen_meals: set[str] = set()

    for row_idx in range(2, ws.max_row + 1):
        meal_value = ws.cell(
            row=row_idx, column=meal_col).value if meal_col else None
        if value_is_missing(meal_value):
            continue

        meal_name = normalize_lookup_label(str(meal_value))
        if meal_name not in distribution:
            issues.append(
                make_issue(
                    section=SECTION_PLAN,
                    message="La comida indicada no es valida.",
                    sheet=PLAN_TEMPLATE_SHEET,
                    location=f"A{row_idx}",
                    field="COMIDA",
                    expected=", ".join(meal_def["name"]
                                       for meal_def in MEAL_DEFS),
                    actual_value=format_actual_value(meal_value),
                )
            )
            continue

        if meal_name in seen_meals:
            issues.append(
                make_issue(
                    section=SECTION_PLAN,
                    message=f"La comida {meal_name} esta repetida.",
                    sheet=PLAN_TEMPLATE_SHEET,
                    location=f"A{row_idx}",
                    field="COMIDA",
                    actual_value=meal_name,
                )
            )
            continue

        for header_name, group_code in EXAMPLE_GROUP_HEADERS.items():
            col_idx = headers.get(header_name)
            if col_idx is None:
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            distribution[meal_name][group_code] = parse_number_with_issue(
                cell.value,
                section=SECTION_PLAN,
                sheet=PLAN_TEMPLATE_SHEET,
                location=cell.coordinate,
                field=header_name,
                issues=issues,
            )

        seen_meals.add(meal_name)

    return distribution


def append_missing_label_issue(
    issues: List[ValidationIssue],
    *,
    section: str,
    sheet: str,
    label_type: str,
    missing_labels: List[str],
) -> None:
    if not missing_labels:
        return
    issues.append(
        make_issue(
            section=section,
            message=f"Faltan etiquetas obligatorias en {label_type}.",
            sheet=sheet,
            field=label_type,
            expected=", ".join(missing_labels),
            actual_value="faltantes",
        )
    )


def build_label_lookup_first(
    rows: List[Tuple[str, object]]
) -> Dict[str, object]:
    lookup: Dict[str, object] = {}
    for label, value in rows:
        normalized = normalize_lookup_label(label)
        if normalized and normalized not in lookup:
            lookup[normalized] = value
    return lookup


def inspect_anthro_data(
    wb,
    patient: PatientInfo,
    issues: List[ValidationIssue],
) -> AnthropometricReportData:
    anthro_data = blank_anthro_data(patient)
    if ANTHRO_TEMPLATE_SHEET not in wb.sheetnames:
        issues.append(
            make_issue(
                section=SECTION_ANTHRO,
                message=f"Falta la hoja obligatoria {ANTHRO_TEMPLATE_SHEET}.",
                sheet=ANTHRO_TEMPLATE_SHEET,
                expected="Hoja con columnas SECCION, ETIQUETA, VALOR",
            )
        )
        return anthro_data

    ws = wb[ANTHRO_TEMPLATE_SHEET]
    headers = build_sheet_headers(ws)
    required_headers = ["SECCION", "ETIQUETA"]
    missing_headers = [
        header for header in required_headers if header not in headers]
    if missing_headers:
        issues.append(
            make_issue(
                section=SECTION_ANTHRO,
                message="Faltan columnas obligatorias en la hoja antropometrica.",
                sheet=ANTHRO_TEMPLATE_SHEET,
                field="columnas",
                expected="SECCION, ETIQUETA y al menos una columna de valores desde la columna C",
                actual_value=", ".join(
                    sorted(headers.keys())) or "sin encabezados",
            )
        )

    section_col = headers.get("SECCION")
    label_col = headers.get("ETIQUETA")
    value_columns = detect_anthro_value_columns(ws)
    if not value_columns:
        issues.append(
            make_issue(
                section=SECTION_ANTHRO,
                message="La hoja antropometrica no tiene columnas de valores.",
                sheet=ANTHRO_TEMPLATE_SHEET,
                field="columnas",
                expected="Al menos una columna de valores desde la columna C",
                actual_value="sin columnas de valores",
            )
        )

    summary_rows_raw: List[Tuple[str, List[object]]] = []
    measurement_rows_raw: List[Tuple[str, List[object]]] = []

    for row_idx in range(2, ws.max_row + 1):
        section_value = ws.cell(
            row=row_idx, column=section_col).value if section_col else None
        label_value = ws.cell(
            row=row_idx, column=label_col).value if label_col else None

        if value_is_missing(section_value) and value_is_missing(label_value):
            continue

        if value_is_missing(section_value):
            issues.append(
                make_issue(
                    section=SECTION_ANTHRO,
                    message="La fila antropometrica no indica la seccion.",
                    sheet=ANTHRO_TEMPLATE_SHEET,
                    location=f"A{row_idx}",
                    field="SECCION",
                    expected="RESUMEN o MEDIDAS",
                    actual_value="vacio",
                )
            )
            continue

        if value_is_missing(label_value):
            issues.append(
                make_issue(
                    section=SECTION_ANTHRO,
                    message="La fila antropometrica no indica la etiqueta.",
                    sheet=ANTHRO_TEMPLATE_SHEET,
                    location=f"B{row_idx}",
                    field="ETIQUETA",
                    expected="nombre de la medida",
                    actual_value="vacio",
                )
            )
            continue

        section_name = normalize_lookup_label(str(section_value))
        row_values = [
            ws.cell(row=row_idx, column=col_idx).value
            for col_idx in value_columns
        ]
        row = (str(label_value).strip(), row_values)
        if section_name == "RESUMEN":
            summary_rows_raw.append(row)
        elif section_name in {"MEDIDAS", "MEDIDA"}:
            measurement_rows_raw.append(row)
        else:
            issues.append(
                make_issue(
                    section=SECTION_ANTHRO,
                    message="La seccion de la fila antropometrica no es valida.",
                    sheet=ANTHRO_TEMPLATE_SHEET,
                    location=f"A{row_idx}",
                    field="SECCION",
                    expected="RESUMEN o MEDIDAS",
                    actual_value=format_actual_value(section_value),
                )
            )

    summary_rows_raw, measurement_rows_raw = sanitize_anthro_rows(
        summary_rows_raw,
        measurement_rows_raw,
    )
    summary_lookup = build_label_lookup_first(summary_rows_raw)
    measurement_lookup = build_label_lookup_first(measurement_rows_raw)
    missing_summary_labels = [
        field
        for field, labels in ANTHRO_REQUIRED_SUMMARY_FIELDS
        if not lookup_contains_any_label(summary_lookup, labels)
    ]
    missing_measurement_labels = [
        field
        for field, labels in ANTHRO_REQUIRED_MEASUREMENT_FIELDS
        if not lookup_contains_any_label(measurement_lookup, labels)
    ]
    append_missing_label_issue(
        issues,
        section=SECTION_ANTHRO,
        sheet=ANTHRO_TEMPLATE_SHEET,
        label_type="RESUMEN",
        missing_labels=missing_summary_labels,
    )
    append_missing_label_issue(
        issues,
        section=SECTION_ANTHRO,
        sheet=ANTHRO_TEMPLATE_SHEET,
        label_type="MEDIDAS",
        missing_labels=missing_measurement_labels,
    )

    anthro_data = AnthropometricReportData(
        patient=patient,
        peso_corporal_kg=format_decimal(
            value_from_lookup(summary_lookup, ANTHRO_PESO_LABELS)
        ),
        estatura_m=format_decimal(
            value_from_lookup(measurement_lookup, ANTHRO_TALLA_M_LABELS)
        ),
        masa_magra_kg=format_decimal(
            value_from_lookup(summary_lookup, ANTHRO_MASA_MAGRA_LABELS)
        ),
        masa_grasa_kg=format_decimal(
            value_from_lookup(summary_lookup, ANTHRO_MASA_GRASA_LABELS)
        ),
        pct_grasa_carter=format_decimal(
            value_from_lookup(
                summary_lookup, ANTHRO_PCT_GRASA_CARTER_LABELS)
        ),
        table_resumen=build_display_anthro_rows(summary_rows_raw),
        table_medidas=build_display_anthro_rows(measurement_rows_raw),
    )

    if not anthro_data.peso_corporal_kg:
        issues.append(
            make_issue(
                section=SECTION_ANTHRO,
                message="No se pudo leer el peso corporal del resumen antropometrico.",
                sheet=ANTHRO_TEMPLATE_SHEET,
                field="Peso (Kg)",
                expected="valor numerico",
            )
        )
    if not anthro_data.estatura_m:
        issues.append(
            make_issue(
                section=SECTION_ANTHRO,
                message="No se pudo leer la talla en metros.",
                sheet=ANTHRO_TEMPLATE_SHEET,
                field="Talla (m)",
                expected="valor numerico",
            )
        )
    if not anthro_data.masa_magra_kg:
        issues.append(
            make_issue(
                section=SECTION_ANTHRO,
                message="No se pudo leer Kg de Masa Magra.",
                sheet=ANTHRO_TEMPLATE_SHEET,
                field="Kg de Masa Magra",
                expected="valor numerico",
            )
        )
    if not anthro_data.masa_grasa_kg:
        issues.append(
            make_issue(
                section=SECTION_ANTHRO,
                message="No se pudo leer Kg de Grasa.",
                sheet=ANTHRO_TEMPLATE_SHEET,
                field="Kg de Grasa",
                expected="valor numerico",
            )
        )
    if not anthro_data.pct_grasa_carter:
        issues.append(
            make_issue(
                section=SECTION_ANTHRO,
                message="No se pudo leer % Grasa (Carter 1986).",
                sheet=ANTHRO_TEMPLATE_SHEET,
                field="% Grasa (Carter 1986)",
                expected="valor numerico",
            )
        )

    return anthro_data


def inspect_meal_examples(
    wb,
    meal_distribution: Dict[str, Dict[str, float]],
    issues: List[ValidationIssue],
) -> tuple[Dict[str, str], str]:
    if "EJEMPLOS_COMIDAS" not in wb.sheetnames:
        return {}, "no cargados"

    try:
        return build_meal_example_texts(wb, meal_distribution), "cargados"
    except ValidationError as exc:
        issues.append(
            make_issue(
                section=SECTION_EXAMPLES,
                message="No se pudieron leer los ejemplos de comidas.",
                sheet="EJEMPLOS_COMIDAS",
                expected="Hoja opcional bien formada",
                actual_value=str(exc),
            )
        )
        return {}, "con errores"


def inspect_workbook(wb) -> ParsedWorkbookData:
    issues: List[ValidationIssue] = []
    patient = inspect_patient_info(wb, issues)
    meal_distribution = inspect_plan_distribution(wb, issues)
    meal_totals = calculate_meal_totals(meal_distribution)
    anthro_data = inspect_anthro_data(wb, patient, issues)
    meal_examples, examples_status = inspect_meal_examples(
        wb, meal_distribution, issues
    )
    anthro_data.patient = patient
    return ParsedWorkbookData(
        patient=patient,
        meal_distribution=meal_distribution,
        meal_totals=meal_totals,
        anthro_data=anthro_data,
        meal_examples=meal_examples,
        issues=issues,
        examples_status=examples_status,
    )


def format_validation_summary(issues: List[ValidationIssue]) -> str:
    blocking_count = sum(1 for issue in issues if issue.is_blocking)
    if blocking_count == 0:
        return "El Excel se valido correctamente."
    if blocking_count == 1:
        return "No se puede generar: se encontro 1 error en el Excel."
    return f"No se puede generar: se encontraron {blocking_count} errores en el Excel."


def format_validation_warning_summary(issues: List[ValidationIssue]) -> str:
    ordered = ordered_issues(issues)
    if not ordered:
        return "El Excel se valido correctamente."

    error_count = sum(1 for issue in ordered if issue.is_blocking)
    warning_count = len(ordered) - error_count

    if error_count and warning_count:
        return (
            f"Se generó el documento con {error_count} errores y "
            f"{warning_count} advertencias detectadas en el Excel."
        )
    if error_count == 1:
        return "Se generó el documento con 1 error detectado en el Excel."
    if error_count > 1:
        return f"Se generó el documento con {error_count} errores detectados en el Excel."
    if warning_count == 1:
        return "Se generó el documento con 1 advertencia detectada en el Excel."
    return f"Se generó el documento con {warning_count} advertencias detectadas en el Excel."


def build_issue_detail(issue: ValidationIssue) -> str:
    parts: List[str] = []
    if issue.sheet:
        parts.append(f"Hoja: {issue.sheet}")
    if issue.location:
        parts.append(f"Ubicacion: {issue.location}")
    if issue.field:
        parts.append(f"Campo: {issue.field}")
    if issue.expected:
        parts.append(f"Esperado: {issue.expected}")
    if issue.actual_value:
        parts.append(f"Encontrado: {issue.actual_value}")
    return " | ".join(parts)


def ordered_issues(issues: List[ValidationIssue]) -> List[ValidationIssue]:
    section_index = {
        section: idx for idx, section in enumerate(SECTION_ORDER)
    }
    unique: dict[tuple[str, str, str, str,
                       str, str, str], ValidationIssue] = {}
    for issue in issues:
        key = (
            issue.section,
            issue.message,
            issue.sheet,
            issue.location,
            issue.field,
            issue.expected,
            issue.actual_value,
        )
        unique.setdefault(key, issue)
    return sorted(
        unique.values(),
        key=lambda issue: (
            section_index.get(issue.section, len(section_index)),
            issue.sheet,
            issue.location,
            issue.message,
        ),
    )


def format_validation_report(issues: List[ValidationIssue]) -> str:
    ordered = ordered_issues(issues)
    if not ordered:
        return "Sin errores detectados."

    lines: List[str] = []
    current_section = ""
    counter = 1
    for issue in ordered:
        if issue.section != current_section:
            if lines:
                lines.append("")
            current_section = issue.section
            lines.append(current_section)
        lines.append(f"{counter}. {issue.message}")
        detail = build_issue_detail(issue)
        if detail:
            lines.append(f"   {detail}")
        counter += 1
    return "\n".join(lines)


def build_validation_error_message(issues: List[ValidationIssue]) -> str:
    return format_validation_summary(issues) + "\n\n" + format_validation_report(issues)


def build_validation_warning_message(issues: List[ValidationIssue]) -> str:
    return format_validation_warning_summary(issues) + "\n\n" + format_validation_report(issues)


def format_preview_text(data: ParsedWorkbookData) -> str:
    lines: List[str] = []
    lines.append("Paciente")
    lines.append(f"- Nombre: {data.patient.name or 'vacio'}")
    lines.append(f"- Cedula: {data.patient.ci or 'vacio'}")
    lines.append(f"- Sexo: {data.patient.sex or 'vacio'}")
    lines.append(f"- Edad: {data.patient.age or 'vacio'}")
    lines.append(f"- Disciplina: {data.patient.discipline or 'vacio'}")

    lines.append("")
    lines.append("Plan leido")
    for meal_def in MEAL_DEFS:
        meal_name = meal_def["name"]
        values = data.meal_distribution.get(meal_name, {})
        rendered = ", ".join(
            f"{GROUP_NAMES[group]}={format_quantity(values.get(group, 0.0))}"
            for group in GROUP_ROWS
        )
        lines.append(f"- {meal_name}: {rendered}")

    lines.append("")
    lines.append("Totales del dia")
    lines.append(
        "- "
        + ", ".join(
            f"{GROUP_NAMES[group]}={format_quantity(data.meal_totals.get(group, 0.0))}"
            for group in GROUP_ROWS
        )
    )

    lines.append("")
    lines.append("Antropometria leida")
    lines.append(f"- Peso: {data.anthro_data.peso_corporal_kg or 'vacio'}")
    lines.append(f"- Talla: {data.anthro_data.estatura_m or 'vacio'}")
    lines.append(f"- Kg masa magra: {data.anthro_data.masa_magra_kg or 'vacio'}")
    lines.append(f"- Kg grasa: {data.anthro_data.masa_grasa_kg or 'vacio'}")
    lines.append(
        f"- % grasa Carter: {data.anthro_data.pct_grasa_carter or 'vacio'}")

    lines.append("")
    lines.append("Resumen antropometrico")
    if data.anthro_data.table_resumen:
        for row in data.anthro_data.table_resumen:
            label = row[0]
            values = [value or "vacio" for value in row[1:]]
            lines.append(f"- {label}: {' | '.join(values) if values else 'vacio'}")
    else:
        lines.append("- Sin filas leidas")

    lines.append("")
    lines.append("Medidas antropometricas")
    if data.anthro_data.table_medidas:
        for row in data.anthro_data.table_medidas:
            label = row[0]
            values = [value or "vacio" for value in row[1:]]
            lines.append(f"- {label}: {' | '.join(values) if values else 'vacio'}")
    else:
        lines.append("- Sin filas leidas")

    lines.append("")
    lines.append(f"Ejemplos: {data.examples_status}")
    if data.meal_examples:
        for meal_name in sorted(data.meal_examples):
            lines.append(f"- {meal_name}: {data.meal_examples[meal_name]}")

    lines.append("")
    lines.append("Errores detectados")
    if data.issues:
        lines.append(format_validation_report(data.issues))
    else:
        lines.append("Sin errores bloqueantes.")

    return "\n".join(lines)
