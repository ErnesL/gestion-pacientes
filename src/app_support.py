from __future__ import annotations

import re
import sys
import unicodedata
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from excel_helpers import ValidationError
from generate_anthro_pptx import generate_anthro_pptx
from generate_pptx import generate_plan_pptx
from pdf_export import export_pptx_to_pdf


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PLAN_TEMPLATE_NAME = "Plan de Alimentación Base.pptx"
ANTHRO_TEMPLATE_NAME = "Informe Antropométrico base.pptx"


@dataclass
class GeneratedDocument:
    label: str
    pptx_path: Path | None = None
    pdf_path: Path | None = None
    errors: list[str] = field(default_factory=list)


@dataclass
class GenerationResult:
    patient_name: str
    documents: list[GeneratedDocument]

    @property
    def warnings(self) -> list[str]:
        warnings: list[str] = []
        for document in self.documents:
            warnings.extend(document.errors)
        return warnings

    @property
    def has_successes(self) -> bool:
        return any(document.pptx_path is not None for document in self.documents)

    @property
    def has_errors(self) -> bool:
        return any(document.errors for document in self.documents)


def resolve_app_root() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return PROJECT_ROOT


def resolve_template_dir() -> Path:
    app_root = resolve_app_root()
    external_dir = app_root / "templates"
    if external_dir.exists():
        return external_dir
    return PROJECT_ROOT / "src-material"


def get_template_paths() -> dict[str, Path]:
    template_dir = resolve_template_dir()
    return {
        "plan": template_dir / PLAN_TEMPLATE_NAME,
        "anthro": template_dir / ANTHRO_TEMPLATE_NAME,
    }


def sanitize_filename_component(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_value = re.sub(r'[<>:"/\\\\|?*]', " ", ascii_value)
    ascii_value = re.sub(r"\s+", " ", ascii_value).strip()
    ascii_value = ascii_value.strip(". ")
    return ascii_value or "Paciente"


def unique_stem(output_dir: Path, stem: str, suffixes: list[str]) -> str:
    candidate = stem
    counter = 2
    while any((output_dir / f"{candidate}{suffix}").exists() for suffix in suffixes):
        candidate = f"{stem} ({counter})"
        counter += 1
    return candidate


def load_patient_name(excel_path: Path) -> str:
    wb = load_workbook(excel_path, data_only=True)
    if "HISTORIA" not in wb.sheetnames:
        raise ValidationError("No existe la hoja requerida: HISTORIA")
    name = str(wb["HISTORIA"]["C4"].value or "").strip()
    if not name:
        raise ValidationError("Falta campo: Nombre y Apellido (HISTORIA!C4)")
    return name


def build_output_stems(output_dir: Path, patient_name: str) -> dict[str, str]:
    safe_name = sanitize_filename_component(patient_name)
    plan_stem = unique_stem(
        output_dir,
        f"Plan Alimentacion - {safe_name}",
        [".pptx", ".pdf"],
    )
    anthro_stem = unique_stem(
        output_dir,
        f"Informe Antropometrico - {safe_name}",
        [".pptx", ".pdf"],
    )
    return {"plan": plan_stem, "anthro": anthro_stem}


def validate_inputs(excel_path: Path, output_dir: Path) -> None:
    if not excel_path.exists():
        raise FileNotFoundError(f"No existe el archivo Excel: {excel_path}")
    if not output_dir.exists():
        raise FileNotFoundError(f"No existe la carpeta destino: {output_dir}")
    if not output_dir.is_dir():
        raise NotADirectoryError(
            f"La ruta destino no es una carpeta: {output_dir}")


def generate_all_documents(
    excel_path: Path | str,
    output_dir: Path | str,
    log: Callable[[str], None],
    today: date | None = None,
) -> GenerationResult:
    excel_path = Path(excel_path)
    output_dir = Path(output_dir)
    template_paths = get_template_paths()

    log("Validando rutas y templates")
    validate_inputs(excel_path, output_dir)

    patient_name = load_patient_name(excel_path)
    stems = build_output_stems(output_dir, patient_name)
    documents = [
        GeneratedDocument(label="Plan de Alimentación"),
        GeneratedDocument(label="Informe Antropométrico"),
    ]

    plan_doc = documents[0]
    plan_doc.pptx_path = output_dir / f"{stems['plan']}.pptx"
    plan_doc.pdf_path = output_dir / f"{stems['plan']}.pdf"

    anthro_doc = documents[1]
    anthro_doc.pptx_path = output_dir / f"{stems['anthro']}.pptx"
    anthro_doc.pdf_path = output_dir / f"{stems['anthro']}.pdf"

    try:
        log("Generando plan de alimentación")
        if not template_paths["plan"].exists():
            raise FileNotFoundError(
                f"No existe el template plan: {template_paths['plan']}")
        generate_plan_pptx(
            excel_path=excel_path,
            template_path=template_paths["plan"],
            output_path=plan_doc.pptx_path,
        )
    except Exception as exc:
        plan_doc.errors.append(f"No se pudo generar el PPTX del plan: {exc}")
        plan_doc.pptx_path = None
        plan_doc.pdf_path = None

    try:
        log("Generando informe antropométrico")
        if not template_paths["anthro"].exists():
            raise FileNotFoundError(
                f"No existe el template antropometrico: {template_paths['anthro']}"
            )
        generate_anthro_pptx(
            excel_path=excel_path,
            template_path=template_paths["anthro"],
            output_path=anthro_doc.pptx_path,
            today=today,
        )
    except Exception as exc:
        anthro_doc.errors.append(
            f"No se pudo generar el PPTX del informe: {exc}")
        anthro_doc.pptx_path = None
        anthro_doc.pdf_path = None

    if plan_doc.pptx_path is not None and plan_doc.pdf_path is not None:
        try:
            log("Exportando PDF del plan")
            export_pptx_to_pdf(plan_doc.pptx_path, plan_doc.pdf_path)
        except Exception as exc:
            plan_doc.errors.append(
                f"No se pudo exportar el PDF del plan: {exc}")
            plan_doc.pdf_path = None

    if anthro_doc.pptx_path is not None and anthro_doc.pdf_path is not None:
        try:
            log("Exportando PDF del informe")
            export_pptx_to_pdf(anthro_doc.pptx_path, anthro_doc.pdf_path)
        except Exception as exc:
            anthro_doc.errors.append(
                f"No se pudo exportar el PDF del informe: {exc}")
            anthro_doc.pdf_path = None

    if not any(document.pptx_path is not None for document in documents):
        all_errors = [
            error for document in documents for error in document.errors]
        if all_errors:
            raise ValidationError("\n".join(all_errors))
        raise ValidationError("No se generó ningún archivo.")

    return GenerationResult(patient_name=patient_name, documents=documents)
