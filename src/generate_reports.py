from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from weasyprint import HTML


PROJECT_ROOT = Path(__file__).resolve().parents[1]
TEMPLATES_DIR = Path(__file__).resolve().parent / "templates"
STYLES_DIR = Path(__file__).resolve().parent / "styles"


@dataclass
class PatientInfo:
    name: str
    ci: str
    sex: str
    age: str
    discipline: str


@dataclass
class PlanData:
    patient: PatientInfo
    agua_litros: str
    distribution: Dict[str, Dict[str, float]]
    totals: Dict[str, float]


@dataclass
class AntropometricoData:
    patient: PatientInfo
    eval_date: str
    weight_kg: float
    height_cm: float
    height_m: float
    bmi: float
    bmi_class: str
    body_fat_pct: float
    lean_mass_kg: float
    fat_mass_kg: float
    somatotype: str
    measurements: List[Dict[str, str]]
    next_control: str


class ValidationError(Exception):
    pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Genera PDFs desde Excel")
    parser.add_argument("excel", help="Ruta al archivo Excel")
    parser.add_argument(
        "--output",
        default=str(PROJECT_ROOT / "output"),
        help="Carpeta de salida (default: output)",
    )
    return parser.parse_args()


def sanitize_filename(value: str) -> str:
    cleaned = "".join(ch for ch in value if ch.isalnum()
                      or ch in " -_").strip()
    return cleaned or "paciente"


def excel_date_to_str(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date().strftime("%d/%m/%Y")
        except Exception:
            return ""
    return ""


def to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(",", "."))
        except ValueError:
            return 0.0
    return 0.0


def format_number(value: float, decimals: int = 2) -> str:
    return f"{value:.{decimals}f}"


def bmi_classification(bmi: float) -> str:
    if bmi <= 0:
        return ""
    if bmi < 18.5:
        return "Bajo peso"
    if bmi < 25:
        return "Normal"
    if bmi < 30:
        return "Sobrepeso I"
    if bmi < 35:
        return "Obesidad I"
    if bmi < 40:
        return "Obesidad II"
    return "Obesidad III"


def load_patient_info(wb) -> PatientInfo:
    ws = wb["HISTORIA"]
    name = str(ws["C4"].value or "").strip()
    ci = str(ws["C5"].value or "").strip()
    age = ws["C7"].value
    sex = str(ws["C10"].value or "").strip()
    discipline = str(ws["I8"].value or "").strip()

    age_str = ""
    if isinstance(age, (int, float)):
        age_str = str(int(round(age)))
    elif age:
        age_str = str(age)

    missing = []
    if not name:
        missing.append("Nombre y Apellido (HISTORIA!C4)")
    if not ci:
        missing.append("Cédula (HISTORIA!C5)")
    if not sex:
        missing.append("Sexo (HISTORIA!C10)")
    if not age_str:
        missing.append("Edad (HISTORIA!C7)")
    if not discipline:
        missing.append("Disciplina/Actividad (HISTORIA!I8)")

    if missing:
        raise ValidationError("Faltan campos: " + "; ".join(missing))

    return PatientInfo(
        name=name,
        ci=str(ci),
        sex=sex,
        age=age_str,
        discipline=discipline,
    )


def load_plan_data(wb, patient: PatientInfo) -> PlanData:
    ws = wb["REQUERIMIENTOS"]

    meal_cols = {
        "Pre Desayuno": "K",
        "Desayuno": "L",
        "Merienda AM": "M",
        "Almuerzo": "N",
        "Merienda PM": "P",
        "Cena": "R",
    }
    group_rows = {
        "Lácteos": 48,
        "Vegetales": 49,
        "Frutas": 50,
        "Almidones": 51,
        "Proteínas": 53,
        "Grasas": 54,
    }

    distribution: Dict[str, Dict[str, float]] = {
        group: {} for group in group_rows}
    totals: Dict[str, float] = {}

    for group, row in group_rows.items():
        for meal, col in meal_cols.items():
            cell = f"{col}{row}"
            distribution[group][meal] = to_float(ws[cell].value)
        total_cell = f"S{row}"
        totals[group] = to_float(ws[total_cell].value)

    agua_litros = format_number(to_float(ws["I39"].value), 2)

    return PlanData(
        patient=patient,
        agua_litros=agua_litros,
        distribution=distribution,
        totals=totals,
    )


def load_antropometrico_data(wb, patient: PatientInfo) -> AntropometricoData:
    ws = wb["RESUMEN ANTROPOMETRICO"]

    eval_date = excel_date_to_str(
        ws["F5"].value) or excel_date_to_str(ws["F36"].value)
    weight_kg = to_float(ws["F6"].value)
    height_cm = to_float(ws["F7"].value)
    height_m = height_cm / 100 if height_cm else 0
    bmi = weight_kg / (height_m ** 2) if height_m else 0
    body_fat_pct = to_float(ws["F8"].value)
    lean_mass_kg = to_float(ws["F11"].value)
    fat_mass_kg = to_float(ws["F12"].value)
    somatotype = str(ws["F16"].value or "").strip()

    measurements = []
    for row in range(36, 62):
        label = ws[f"E{row}"].value
        value = ws[f"F{row}"].value
        if not label:
            continue
        if isinstance(value, str) and value.strip().startswith("#"):
            value_str = ""
        else:
            value_str = format_number(
                to_float(value), 2) if value is not None else ""
        measurements.append({"label": str(label).strip(), "value": value_str})

    next_control_date = (date.today() + timedelta(weeks=6)
                         ).strftime("%d/%m/%Y")

    return AntropometricoData(
        patient=patient,
        eval_date=eval_date,
        weight_kg=weight_kg,
        height_cm=height_cm,
        height_m=height_m,
        bmi=bmi,
        bmi_class=bmi_classification(bmi),
        body_fat_pct=body_fat_pct,
        lean_mass_kg=lean_mass_kg,
        fat_mass_kg=fat_mass_kg,
        somatotype=somatotype,
        measurements=measurements,
        next_control=next_control_date,
    )


def render_pdf(template_name: str, stylesheet: Path, context: Dict[str, Any], output_path: Path) -> None:
    env = Environment(
        loader=FileSystemLoader(str(TEMPLATES_DIR)),
        autoescape=select_autoescape(["html"]),
    )
    template = env.get_template(template_name)
    html_content = template.render(**context)

    html = HTML(string=html_content, base_url=str(PROJECT_ROOT))
    html.write_pdf(
        target=str(output_path),
        stylesheets=[str(stylesheet)],
    )


def main() -> int:
    args = parse_args()
    excel_path = Path(args.excel)
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    if not excel_path.exists():
        print(f"No existe el archivo: {excel_path}")
        return 1

    wb = load_workbook(excel_path, data_only=True)

    try:
        patient = load_patient_info(wb)
        plan = load_plan_data(wb, patient)
        antro = load_antropometrico_data(wb, patient)
    except ValidationError as exc:
        print(f"Error: {exc}")
        return 1

    safe_name = sanitize_filename(patient.name)

    plan_output = output_dir / f"Plan Alimentacion - {safe_name}.pdf"
    antro_output = output_dir / f"Informe Antropometrico - {safe_name}.pdf"

    render_pdf(
        "plan.html",
        STYLES_DIR / "plan.css",
        {"data": plan},
        plan_output,
    )

    render_pdf(
        "antropometrico.html",
        STYLES_DIR / "antropometrico.css",
        {"data": antro},
        antro_output,
    )

    print(f"Generados: {plan_output} y {antro_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
