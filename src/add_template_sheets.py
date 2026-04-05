from __future__ import annotations

import argparse
import ast
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
import re
from typing import Iterable, Sequence

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from excel_helpers import (
    ANTHRO_REQUIRED_MEASUREMENT_FIELDS,
    ANTHRO_TEMPLATE_SHEET,
    GROUP_ROWS,
    MEAL_DEFS,
    PLAN_TEMPLATE_SHEET,
)


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE_WORKBOOK = (
    PROJECT_ROOT / "examples" / "ejemplo-config-ejemplos-comidas.xlsx"
)
DEFAULT_SHEETS_TO_ADD = (
    "PLAN_ALIMENTACION_TEMPLATE",
    "ANTROPOMETRIA_TEMPLATE",
    "EJEMPLOS_COMIDAS",
    "EQUIVALENCIAS_EJEMPLOS",
)
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
LEGACY_SUMMARY_SHEET = "RESUMEN ANTROPOMETRICO"
LEGACY_ANTHRO_SHEET = "ANTROPOMETRÍA DEPORTISTAS"
LEGACY_PLAN_SHEET = "REQUERIMIENTOS"
DEFAULT_COLUMN_WIDTH = 14.0
COMPACT_ROW_HEIGHT = 18.0
EXAMPLES_SHEET = "EJEMPLOS_COMIDAS"
EXAMPLE_EQUIVALENCES_SHEET = "EQUIVALENCIAS_EJEMPLOS"
PLAN_TEMPLATE_HEADERS = [
    "COMIDA",
    "LACTEOS",
    "VEGETALES",
    "FRUTAS",
    "ALMIDONES",
    "PROTEINAS",
    "GRASAS",
]
EXAMPLES_TEMPLATE_HEADERS = [
    "COMIDA",
    "LACTEOS",
    "VEGETALES",
    "FRUTAS",
    "ALMIDONES",
    "PROTEINAS",
    "GRASAS",
    "OBSERVACION",
]
EQUIVALENCE_TEMPLATE_HEADERS = [
    "CODIGO ALIMENTO",
    "GRUPO",
    "DESCRIPCION BASE",
    "CANTIDAD POR RACION",
    "USA DECIMAL",
    "TEXTO SINGULAR",
    "TEXTO PLURAL",
]
PLAN_TEMPLATE_COLUMN_WIDTHS = {
    "A": 12.0,
    "B": 11.0,
    "C": 11.0,
    "D": 11.0,
    "E": 12.0,
    "F": 12.0,
    "G": 11.0,
}
EXAMPLES_TEMPLATE_COLUMN_WIDTHS = {
    "A": 12.0,
    "B": 20.0,
    "C": 20.0,
    "D": 16.0,
    "E": 16.0,
    "F": 18.0,
    "G": 16.0,
    "H": 16.0,
}
EQUIVALENCE_TEMPLATE_COLUMN_WIDTHS = {
    "A": 22.0,
    "B": 14.0,
    "C": 22.0,
    "D": 16.0,
    "E": 12.0,
    "F": 12.0,
    "G": 28.0,
    "H": 28.0,
    "I": 24.0,
}

LEGACY_SUMMARY_BACKFILL_FIELDS = [
    ("Evaluación", ["Evaluación", "Número de evaluación", "Numero de evaluación"]),
    ("Fecha", ["Fecha", "Fecha de evaluación"]),
    ("Peso (Kg)", ["Peso (Kg)", "Peso actual (kg)", "Peso Actual (Kg)"]),
    ("Talla parada (cm)", ["Talla parada (cm)", "Talla (cm)"]),
    ("% Grasa (Carter 1986)", ["% Grasa (Carter 1986)", "% Grasa Carter", "%grasa carter"]),
    ("% Grasa (Durnin y W. 1974)", ["% Grasa (Durnin y W. 1974)"]),
    ("Interpretación", ["Interpretación"]),
    ("Kg de Masa Magra", ["Kg de Masa Magra", "Kg. Masa magra"]),
    ("Kg de Grasa", ["Kg de Grasa", "Kg. Grasa"]),
    ("Masa Muscular (Kg)", ["Masa Muscular (Kg)"]),
    ("Masa Adiposa (Kg)", ["Masa Adiposa (Kg)"]),
    ("Sumatoria de 6 pliegues", ["Sumatoria de 6 pliegues", "Sumatoria de 6 Pliegues"]),
    ("Somatotipo", ["Somatotipo", "SOMATOTIPO"]),
]

LEGACY_MEASUREMENTS_BACKFILL_FIELDS = [
    (
        field,
        labels + (["Brazo Flexionado en Tensón (cm)"] if field == "Brazo Flexionado en Tensión (cm)" else []),
    )
    for field, labels in ANTHRO_REQUIRED_MEASUREMENT_FIELDS
]
BACKFILL_REQUIRED_SUMMARY_LABELS = {
    "FECHA",
    "PESO (KG)",
    "TALLA PARADA (CM)",
    "% GRASA (CARTER 1986)",
    "KG DE GRASA",
}
BACKFILL_REQUIRED_MEASUREMENT_LABELS = {
    "FECHA DE EVALUACION",
    "TALLA (M)",
}
BACKFILL_DATE_LABELS = {
    "FECHA",
    "FECHA DE EVALUACION",
}
BACKFILL_POSITIVE_NUMBER_LABELS = (
    BACKFILL_REQUIRED_SUMMARY_LABELS | BACKFILL_REQUIRED_MEASUREMENT_LABELS
) - BACKFILL_DATE_LABELS
FORMULA_EXACT_REFERENCE_PATTERN = re.compile(
    r"^\s*(?:(?P<sheet>'(?:[^']|'')+'|[A-Za-z0-9_ .ÁÉÍÓÚÜÑ()\-]+)!)?(?P<coord>\$?[A-Z]{1,3}\$?\d+)\s*$"
)
FORMULA_CELL_REFERENCE_PATTERN = re.compile(
    r"(?:(?P<sheet>'(?:[^']|'')+'|[A-Za-z0-9_ .ÁÉÍÓÚÜÑ()\-]+)!)?(?P<coord>\$?[A-Z]{1,3}\$?\d+)"
)


@dataclass
class WorkbookFormulaResolver:
    value_wb: Workbook
    formula_wb: Workbook
    cache: dict[tuple[str, str], object] = field(default_factory=dict)

    def has_source_content(self, sheet_name: str, row_idx: int, col_idx: int) -> bool:
        if sheet_name in self.value_wb.sheetnames:
            value = self.value_wb[sheet_name].cell(row=row_idx, column=col_idx).value
            if not value_is_missing(value):
                return True
        if sheet_name not in self.formula_wb.sheetnames:
            return False
        raw_value = self.formula_wb[sheet_name].cell(row=row_idx, column=col_idx).value
        return not value_is_missing(raw_value)

    def resolve_cell(self, sheet_name: str, row_idx: int, col_idx: int) -> object:
        return self.resolve_coordinate(
            sheet_name,
            f"{get_column_letter(col_idx)}{row_idx}",
        )

    def resolve_coordinate(
        self,
        sheet_name: str,
        coord: str,
        *,
        visited: set[tuple[str, str]] | None = None,
    ) -> object:
        normalized_coord = coord.replace("$", "")
        cache_key = (sheet_name, normalized_coord)
        if cache_key in self.cache:
            return self.cache[cache_key]
        if visited is None:
            visited = set()
        if cache_key in visited:
            return None
        visited = set(visited)
        visited.add(cache_key)

        value_ws = self.value_wb[sheet_name] if sheet_name in self.value_wb.sheetnames else None
        raw_ws = self.formula_wb[sheet_name] if sheet_name in self.formula_wb.sheetnames else None
        if value_ws is not None:
            cached_value = value_ws[normalized_coord].value
            if not value_is_missing(cached_value):
                self.cache[cache_key] = cached_value
                return cached_value
        if raw_ws is None:
            self.cache[cache_key] = None
            return None

        raw_value = raw_ws[normalized_coord].value
        if value_is_missing(raw_value):
            self.cache[cache_key] = None
            return None
        if not (isinstance(raw_value, str) and raw_value.startswith("=")):
            self.cache[cache_key] = raw_value
            return raw_value

        resolved_value = self._evaluate_formula(
            raw_value[1:],
            current_sheet_name=sheet_name,
            visited=visited,
        )
        self.cache[cache_key] = resolved_value
        return resolved_value

    def _evaluate_formula(
        self,
        expression: str,
        *,
        current_sheet_name: str,
        visited: set[tuple[str, str]],
    ) -> object:
        exact_match = FORMULA_EXACT_REFERENCE_PATTERN.fullmatch(expression.strip())
        if exact_match:
            target_sheet = self._normalize_formula_sheet_name(
                exact_match.group("sheet"),
                default_sheet=current_sheet_name,
            )
            return self.resolve_coordinate(
                target_sheet,
                exact_match.group("coord"),
                visited=visited,
            )

        def replace_reference(match: re.Match[str]) -> str:
            target_sheet = self._normalize_formula_sheet_name(
                match.group("sheet"),
                default_sheet=current_sheet_name,
            )
            resolved = self.resolve_coordinate(
                target_sheet,
                match.group("coord"),
                visited=visited,
            )
            parsed = parse_backfill_number(resolved)
            if parsed is None:
                raise ValueError("Referencia no numerica")
            return repr(parsed)

        try:
            python_expression = FORMULA_CELL_REFERENCE_PATTERN.sub(
                replace_reference,
                expression,
            ).replace("^", "**")
            node = ast.parse(python_expression, mode="eval")
            if not self._is_safe_numeric_ast(node):
                return None
            return eval(compile(node, "<formula>", "eval"), {"__builtins__": {}}, {})
        except Exception:
            return None

    def _normalize_formula_sheet_name(
        self,
        raw_sheet_name: str | None,
        *,
        default_sheet: str,
    ) -> str:
        if raw_sheet_name is None:
            return default_sheet
        if raw_sheet_name.startswith("'") and raw_sheet_name.endswith("'"):
            return raw_sheet_name[1:-1].replace("''", "'")
        return raw_sheet_name

    def _is_safe_numeric_ast(self, node: ast.AST) -> bool:
        if isinstance(node, ast.Expression):
            return self._is_safe_numeric_ast(node.body)
        if isinstance(node, ast.Constant):
            return isinstance(node.value, (int, float))
        if isinstance(node, ast.UnaryOp):
            return isinstance(node.op, (ast.UAdd, ast.USub)) and self._is_safe_numeric_ast(node.operand)
        if isinstance(node, ast.BinOp):
            return (
                isinstance(node.op, (ast.Add, ast.Sub, ast.Mult, ast.Div, ast.Pow))
                and self._is_safe_numeric_ast(node.left)
                and self._is_safe_numeric_ast(node.right)
            )
        return False


@dataclass(frozen=True)
class WorkbookSheetUpdateResult:
    path: Path
    added_sheets: tuple[str, ...]
    replaced_sheets: tuple[str, ...]
    skipped_sheets: tuple[str, ...]

    @property
    def changed(self) -> bool:
        return bool(self.added_sheets or self.replaced_sheets)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Actualiza recursivamente los Excel de una carpeta con nuestras hojas template."
    )
    parser.add_argument("folder", help="Carpeta con archivos .xlsx o .xlsm")
    return parser.parse_args()


def selected_sheet_names() -> tuple[str, ...]:
    return DEFAULT_SHEETS_TO_ADD


def workbook_needs_vba(path: Path) -> bool:
    return path.suffix.lower() == ".xlsm"


def load_excel_workbook(path: Path, *, data_only: bool = False) -> Workbook:
    return load_workbook(
        path,
        keep_vba=workbook_needs_vba(path),
        data_only=data_only,
    )


def iter_excel_files(folder: Path, recursive: bool = False) -> Iterable[Path]:
    iterator = folder.rglob("*") if recursive else folder.iterdir()
    for path in sorted(iterator):
        if not path.is_file():
            continue
        if path.name.startswith("~$"):
            continue
        if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue
        yield path


def value_is_missing(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def normalize_lookup_label(value: str) -> str:
    normalized = value.replace("_", " ").strip().upper()
    for source, target in (
        ("Á", "A"),
        ("É", "E"),
        ("Í", "I"),
        ("Ó", "O"),
        ("Ú", "U"),
        ("Ü", "U"),
    ):
        normalized = normalized.replace(source, target)
    return " ".join(normalized.split())


def detect_last_used_col(
    ws,
    start_col: int = 3,
    *,
    row_start: int = 1,
    row_end: int | None = None,
    formula_resolver: WorkbookFormulaResolver | None = None,
) -> int:
    if row_end is None:
        row_end = ws.max_row
    last_used_col = 0
    for col_idx in range(start_col, ws.max_column + 1):
        if any(
            (
                formula_resolver.has_source_content(ws.title, row_idx, col_idx)
                if formula_resolver is not None
                else not value_is_missing(ws.cell(row=row_idx, column=col_idx).value)
            )
            for row_idx in range(row_start, row_end + 1)
        ):
            last_used_col = col_idx
    return last_used_col


def normalize_backfill_rows(
    rows: list[tuple[str, list[object]]],
) -> list[tuple[str, list[object]]]:
    if not rows:
        return []

    max_len = max(len(values) for _, values in rows)
    if max_len == 0:
        return rows

    last_used_idx = 0
    for idx in range(max_len):
        if any(
            idx < len(values) and not value_is_missing(values[idx])
            for _, values in rows
        ):
            last_used_idx = idx + 1

    if last_used_idx == 0:
        last_used_idx = max_len

    normalized_rows: list[tuple[str, list[object]]] = []
    for label, values in rows:
        trimmed = list(values[:last_used_idx])
        trimmed.extend([None] * (last_used_idx - len(trimmed)))
        normalized_rows.append((label, trimmed))
    return normalized_rows


def build_row_lookup(
    ws,
    label_col: int,
    *,
    row_start: int = 1,
    row_end: int | None = None,
) -> dict[str, int]:
    if row_end is None:
        row_end = ws.max_row
    lookup: dict[str, int] = {}
    for row_idx in range(row_start, row_end + 1):
        label_value = ws.cell(row=row_idx, column=label_col).value
        if value_is_missing(label_value):
            continue
        normalized = normalize_lookup_label(str(label_value))
        if normalized and normalized not in lookup:
            lookup[normalized] = row_idx
    return lookup


def build_sheet_headers(ws, header_row: int = 1) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[header_row]:
        if value_is_missing(cell.value):
            continue
        headers[normalize_lookup_label(str(cell.value))] = cell.column
    return headers


def find_matching_row(
    row_lookup: dict[str, int],
    aliases: Sequence[str],
) -> int | None:
    for alias in aliases:
        normalized = normalize_lookup_label(alias)
        if normalized in row_lookup:
            return row_lookup[normalized]
    return None


def extract_rows_from_sheet(
    ws,
    *,
    label_col: int,
    value_cols: Sequence[int],
    field_defs: Sequence[tuple[str, Sequence[str]]],
    row_start: int = 1,
    row_end: int | None = None,
    formula_resolver: WorkbookFormulaResolver | None = None,
) -> list[tuple[str, list[object]]]:
    row_lookup = build_row_lookup(
        ws,
        label_col,
        row_start=row_start,
        row_end=row_end,
    )
    rows: list[tuple[str, list[object]]] = []
    for canonical_label, aliases in field_defs:
        row_idx = find_matching_row(row_lookup, aliases)
        if row_idx is None:
            rows.append((canonical_label, [None] * len(value_cols)))
            continue
        values = [
            (
                formula_resolver.resolve_cell(ws.title, row_idx, col_idx)
                if formula_resolver is not None
                else ws.cell(row=row_idx, column=col_idx).value
            )
            for col_idx in value_cols
        ]
        rows.append((canonical_label, values))
    return normalize_backfill_rows(rows)


def rows_have_required_values(
    rows: list[tuple[str, list[object]]],
    required_labels: set[str],
) -> bool:
    found_labels: set[str] = set()
    for label, values in rows:
        normalized = normalize_lookup_label(label)
        if normalized not in required_labels:
            continue
        if any(not value_is_missing(value) for value in values):
            found_labels.add(normalized)
    return required_labels.issubset(found_labels)


def parse_backfill_number(value) -> float | None:
    if value_is_missing(value) or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def backfill_value_is_valid(label: str, value) -> bool:
    normalized = normalize_lookup_label(label)
    if value_is_missing(value):
        return False
    if normalized in BACKFILL_DATE_LABELS:
        return isinstance(value, (datetime, date)) or (
            isinstance(value, str) and bool(value.strip())
        )
    if normalized in BACKFILL_POSITIVE_NUMBER_LABELS:
        parsed = parse_backfill_number(value)
        return parsed is not None and parsed > 0
    return True


def rows_have_required_valid_values(
    rows: list[tuple[str, list[object]]],
    required_labels: set[str],
) -> bool:
    found_labels: set[str] = set()
    for label, values in rows:
        normalized = normalize_lookup_label(label)
        if normalized not in required_labels:
            continue
        if any(backfill_value_is_valid(label, value) for value in values):
            found_labels.add(normalized)
    return required_labels.issubset(found_labels)


def extract_table_rows(
    ws: Worksheet,
    *,
    canonical_headers: Sequence[str],
    key_header: str,
) -> list[list[object]]:
    headers = build_sheet_headers(ws)
    normalized_key = normalize_lookup_label(key_header)
    if normalized_key not in headers:
        return []

    rows: list[list[object]] = []
    key_col = headers[normalized_key]
    for row_idx in range(2, ws.max_row + 1):
        key_value = ws.cell(row=row_idx, column=key_col).value
        if value_is_missing(key_value):
            continue
        rows.append([
            ws.cell(
                row=row_idx,
                column=headers[normalize_lookup_label(header)],
            ).value
            if normalize_lookup_label(header) in headers
            else None
            for header in canonical_headers
        ])
    return rows


def clear_sheet_rows(target_ws: Worksheet, *, start_row: int = 2) -> None:
    if target_ws.max_row >= start_row:
        target_ws.delete_rows(start_row, target_ws.max_row - start_row + 1)


def write_table_rows(target_ws: Worksheet, rows: Sequence[Sequence[object]]) -> None:
    clear_sheet_rows(target_ws, start_row=2)
    for row in rows:
        target_ws.append(list(row))


def hide_empty_columns_after(
    target_ws: Worksheet,
    *,
    last_visible_col: int,
) -> None:
    for col_idx in range(1, target_ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        target_ws.column_dimensions[col_letter].hidden = False

    for col_idx in range(last_visible_col + 1, target_ws.max_column + 1):
        if any(
            not value_is_missing(target_ws.cell(row=row_idx, column=col_idx).value)
            for row_idx in range(1, target_ws.max_row + 1)
        ):
            continue
        target_ws.column_dimensions[get_column_letter(col_idx)].hidden = True


def reset_column_dimensions(target_ws: Worksheet) -> None:
    # Avoid overlapping <col> ranges when the copied template already contains
    # grouped dimensions and we later override widths/hide flags per column.
    target_ws.column_dimensions.clear()


def apply_compact_table_layout(
    target_ws: Worksheet,
    *,
    column_widths: dict[str, float],
    last_visible_col: int,
    row_height: float = COMPACT_ROW_HEIGHT,
) -> None:
    reset_column_dimensions(target_ws)
    target_ws.freeze_panes = "A2"
    for row_idx in range(1, target_ws.max_row + 1):
        target_ws.row_dimensions[row_idx].height = row_height
    for col_letter, width in column_widths.items():
        target_ws.column_dimensions[col_letter].width = width
    hide_empty_columns_after(target_ws, last_visible_col=last_visible_col)


def build_backfill_lookup(
    rows: list[tuple[str, list[object]]],
) -> dict[str, list[object]]:
    lookup: dict[str, list[object]] = {}
    for label, values in rows:
        normalized = normalize_lookup_label(label)
        if normalized and normalized not in lookup:
            lookup[normalized] = list(values)
    return lookup


def filter_rows_by_indices(
    rows: list[tuple[str, list[object]]],
    indices: list[int],
) -> list[tuple[str, list[object]]]:
    return [
        (
            label,
            [values[idx] if idx < len(values) else None for idx in indices],
        )
        for label, values in rows
    ]


def keep_only_valid_backfill_columns(
    summary_rows: list[tuple[str, list[object]]],
    measurement_rows: list[tuple[str, list[object]]],
) -> tuple[list[tuple[str, list[object]]], list[tuple[str, list[object]]]] | None:
    summary_rows = normalize_backfill_rows(summary_rows)
    measurement_rows = normalize_backfill_rows(measurement_rows)
    max_len = max(
        [len(values) for _, values in summary_rows + measurement_rows],
        default=0,
    )
    if max_len == 0:
        return None

    summary_lookup = build_backfill_lookup(summary_rows)
    measurement_lookup = build_backfill_lookup(measurement_rows)
    valid_indices: list[int] = []

    for idx in range(max_len):
        summary_ok = all(
            idx < len(summary_lookup.get(label, []))
            and backfill_value_is_valid(summary_label, summary_lookup[label][idx])
            for summary_label, label in (
                ("Fecha", "FECHA"),
                ("Peso (Kg)", "PESO (KG)"),
                ("Talla parada (cm)", "TALLA PARADA (CM)"),
                ("% Grasa (Carter 1986)", "% GRASA (CARTER 1986)"),
                ("Kg de Grasa", "KG DE GRASA"),
            )
        )
        measurement_ok = all(
            idx < len(measurement_lookup.get(label, []))
            and backfill_value_is_valid(
                measurement_label,
                measurement_lookup[label][idx],
            )
            for measurement_label, label in (
                ("Fecha de evaluación", "FECHA DE EVALUACION"),
                ("Talla (m)", "TALLA (M)"),
            )
        )
        if summary_ok and measurement_ok:
            valid_indices.append(idx)

    if not valid_indices:
        return None

    return (
        filter_rows_by_indices(summary_rows, valid_indices),
        filter_rows_by_indices(measurement_rows, valid_indices),
    )


def render_evaluation_values(values: list[object]) -> list[object]:
    numeric_values: list[int] = []
    for value in values:
        parsed = parse_backfill_number(value)
        if parsed is None or not float(parsed).is_integer():
            return values
        numeric_values.append(int(parsed))
    expected = list(range(1, len(values) + 1))
    if numeric_values != expected:
        return values
    return [
        "1era evaluación" if idx == 0 else f"Control {idx}°"
        for idx in range(len(values))
    ]


def normalize_summary_evaluation_rows(
    summary_rows: list[tuple[str, list[object]]],
) -> list[tuple[str, list[object]]]:
    normalized_rows: list[tuple[str, list[object]]] = []
    for label, values in summary_rows:
        if normalize_lookup_label(label) == "EVALUACION":
            normalized_rows.append((label, render_evaluation_values(list(values))))
            continue
        normalized_rows.append((label, values))
    return normalized_rows


def rows_to_lookup(
    rows: list[tuple[str, list[object]]],
) -> dict[str, list[object]]:
    lookup: dict[str, list[object]] = {}
    for label, values in rows:
        normalized = normalize_lookup_label(label)
        if normalized and normalized not in lookup:
            lookup[normalized] = list(values)
    return lookup


def derive_measurement_rows_from_summary(
    summary_rows: list[tuple[str, list[object]]],
    measurement_rows: list[tuple[str, list[object]]],
) -> list[tuple[str, list[object]]]:
    summary_lookup = rows_to_lookup(summary_rows)
    measurement_lookup = rows_to_lookup(measurement_rows)
    value_count = max(
        [len(values) for _, values in summary_rows + measurement_rows],
        default=0,
    )

    def values_from_lookup(key: str) -> list[object]:
        values = list(summary_lookup.get(key, []))
        values.extend([None] * (value_count - len(values)))
        return values[:value_count]

    def talla_m_values() -> list[object]:
        derived_values: list[object] = []
        for value in values_from_lookup("TALLA PARADA (CM)"):
            parsed = parse_backfill_number(value)
            if parsed is None or parsed <= 0:
                derived_values.append(None)
                continue
            derived_values.append(parsed / 100.0 if parsed > 3 else parsed)
        return derived_values

    fallback_by_label = {
        "FECHA DE EVALUACION": values_from_lookup("FECHA"),
        "PESO ACTUAL (KG)": values_from_lookup("PESO (KG)"),
        "TALLA (M)": talla_m_values(),
        "TALLA (CM)": values_from_lookup("TALLA PARADA (CM)"),
    }

    merged_rows: list[tuple[str, list[object]]] = []
    for label, values in measurement_rows:
        normalized = normalize_lookup_label(label)
        current_values = list(values[:value_count])
        current_values.extend([None] * (value_count - len(current_values)))
        fallback_values = fallback_by_label.get(normalized)
        if fallback_values is None:
            merged_rows.append((label, current_values))
            continue
        merged_rows.append(
            (
                label,
                [
                    current_values[idx]
                    if backfill_value_is_valid(label, current_values[idx])
                    else fallback_values[idx]
                    for idx in range(value_count)
                ],
            )
        )
    return merged_rows


def extract_legacy_summary_layout_rows(
    ws: Worksheet,
    *,
    summary_label_col: int,
    summary_value_start_col: int,
    summary_row_start: int,
    summary_row_end: int | None,
    measurement_label_col: int,
    measurement_value_start_col: int,
    measurement_row_start: int,
    measurement_row_end: int | None,
    formula_resolver: WorkbookFormulaResolver | None = None,
) -> tuple[list[tuple[str, list[object]]], list[tuple[str, list[object]]]] | None:
    summary_last_col = detect_last_used_col(
        ws,
        start_col=summary_value_start_col,
        row_start=summary_row_start,
        row_end=summary_row_end,
        formula_resolver=formula_resolver,
    )
    measurement_last_col = detect_last_used_col(
        ws,
        start_col=measurement_value_start_col,
        row_start=measurement_row_start,
        row_end=measurement_row_end,
        formula_resolver=formula_resolver,
    )
    if summary_last_col < summary_value_start_col or measurement_last_col < measurement_value_start_col:
        return None

    summary_rows = extract_rows_from_sheet(
        ws,
        label_col=summary_label_col,
        value_cols=list(range(summary_value_start_col, summary_last_col + 1)),
        field_defs=LEGACY_SUMMARY_BACKFILL_FIELDS,
        row_start=summary_row_start,
        row_end=summary_row_end,
        formula_resolver=formula_resolver,
    )
    measurement_rows = extract_rows_from_sheet(
        ws,
        label_col=measurement_label_col,
        value_cols=list(range(measurement_value_start_col, measurement_last_col + 1)),
        field_defs=LEGACY_MEASUREMENTS_BACKFILL_FIELDS,
        row_start=measurement_row_start,
        row_end=measurement_row_end,
        formula_resolver=formula_resolver,
    )

    if not rows_have_required_valid_values(summary_rows, BACKFILL_REQUIRED_SUMMARY_LABELS):
        return None
    if not rows_have_required_valid_values(measurement_rows, BACKFILL_REQUIRED_MEASUREMENT_LABELS):
        measurement_rows = derive_measurement_rows_from_summary(
            summary_rows,
            measurement_rows,
        )
    if not rows_have_required_valid_values(measurement_rows, BACKFILL_REQUIRED_MEASUREMENT_LABELS):
        return None

    filtered = keep_only_valid_backfill_columns(summary_rows, measurement_rows)
    if filtered is None:
        return None
    filtered_summary_rows, filtered_measurement_rows = filtered
    return (
        normalize_summary_evaluation_rows(filtered_summary_rows),
        filtered_measurement_rows,
    )


def build_template_backfill_rows(
    source_wb_values: Workbook,
) -> tuple[list[tuple[str, list[object]]], list[tuple[str, list[object]]]] | None:
    if ANTHRO_TEMPLATE_SHEET not in source_wb_values.sheetnames:
        return None

    ws = source_wb_values[ANTHRO_TEMPLATE_SHEET]
    last_used_col = detect_last_used_col(ws, start_col=3)
    if last_used_col < 3:
        return None

    value_cols = list(range(3, last_used_col + 1))
    summary_rows: list[tuple[str, list[object]]] = []
    measurement_rows: list[tuple[str, list[object]]] = []

    for row_idx in range(2, ws.max_row + 1):
        section_value = ws.cell(row=row_idx, column=1).value
        label_value = ws.cell(row=row_idx, column=2).value
        if value_is_missing(section_value) or value_is_missing(label_value):
            continue

        values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in value_cols]
        normalized_section = normalize_lookup_label(str(section_value))
        row = (str(label_value).strip(), values)
        if normalized_section == "RESUMEN":
            summary_rows.append(row)
        elif normalized_section in {"MEDIDAS", "MEDIDA"}:
            measurement_rows.append(row)

    if not summary_rows or not measurement_rows:
        return None

    normalized_summary_rows = normalize_backfill_rows(summary_rows)
    normalized_measurement_rows = normalize_backfill_rows(measurement_rows)
    if not rows_have_required_valid_values(
        normalized_summary_rows, BACKFILL_REQUIRED_SUMMARY_LABELS
    ):
        return None
    if not rows_have_required_valid_values(
        normalized_measurement_rows, BACKFILL_REQUIRED_MEASUREMENT_LABELS
    ):
        return None
    return keep_only_valid_backfill_columns(
        normalized_summary_rows,
        normalized_measurement_rows,
    )


def anthro_rows_equal(
    left: tuple[list[tuple[str, list[object]]], list[tuple[str, list[object]]]] | None,
    right: tuple[list[tuple[str, list[object]]], list[tuple[str, list[object]]]] | None,
) -> bool:
    return left == right


def build_legacy_summary_backfill_rows(
    source_wb_values: Workbook,
    *,
    formula_resolver: WorkbookFormulaResolver | None = None,
) -> tuple[list[tuple[str, list[object]]], list[tuple[str, list[object]]]] | None:
    if LEGACY_SUMMARY_SHEET not in source_wb_values.sheetnames:
        return None

    ws = source_wb_values[LEGACY_SUMMARY_SHEET]
    for layout in (
        {
            "summary_label_col": 2,
            "summary_value_start_col": 4,
            "summary_row_start": 1,
            "summary_row_end": None,
            "measurement_label_col": 2,
            "measurement_value_start_col": 4,
            "measurement_row_start": 1,
            "measurement_row_end": None,
        },
        {
            "summary_label_col": 4,
            "summary_value_start_col": 6,
            "summary_row_start": 16,
            "summary_row_end": 30,
            "measurement_label_col": 5,
            "measurement_value_start_col": 6,
            "measurement_row_start": 34,
            "measurement_row_end": 59,
        },
        ):
        rows = extract_legacy_summary_layout_rows(
            ws,
            **layout,
            formula_resolver=formula_resolver,
        )
        if rows is not None:
            return rows
    return None


def build_legacy_anthro_backfill_rows(
    source_wb_values: Workbook,
    *,
    formula_resolver: WorkbookFormulaResolver | None = None,
) -> tuple[list[tuple[str, list[object]]], list[tuple[str, list[object]]]] | None:
    if LEGACY_ANTHRO_SHEET not in source_wb_values.sheetnames:
        return None

    ws = source_wb_values[LEGACY_ANTHRO_SHEET]
    last_used_col = detect_last_used_col(
        ws,
        start_col=2,
        formula_resolver=formula_resolver,
    )
    if last_used_col < 2:
        return None

    value_cols = list(range(2, last_used_col + 1))
    summary_rows = extract_rows_from_sheet(
        ws,
        label_col=1,
        value_cols=value_cols,
        field_defs=LEGACY_SUMMARY_BACKFILL_FIELDS,
        formula_resolver=formula_resolver,
    )
    measurement_rows = extract_rows_from_sheet(
        ws,
        label_col=1,
        value_cols=value_cols,
        field_defs=LEGACY_MEASUREMENTS_BACKFILL_FIELDS,
        formula_resolver=formula_resolver,
    )

    if not rows_have_required_valid_values(summary_rows, BACKFILL_REQUIRED_SUMMARY_LABELS):
        return None
    if not rows_have_required_valid_values(measurement_rows, BACKFILL_REQUIRED_MEASUREMENT_LABELS):
        return None

    if summary_rows:
        eval_values = summary_rows[0][1]
        if all(value_is_missing(value) for value in eval_values):
            generated_labels = [
                "1era evaluación" if idx == 0 else f"Control {idx}°"
                for idx in range(len(value_cols))
            ]
            summary_rows[0] = ("Evaluación", generated_labels)

    filtered = keep_only_valid_backfill_columns(summary_rows, measurement_rows)
    if filtered is None:
        return None
    filtered_summary_rows, filtered_measurement_rows = filtered
    return (
        normalize_summary_evaluation_rows(filtered_summary_rows),
        filtered_measurement_rows,
    )


def resolve_anthro_backfill_rows(
    source_wb_values: Workbook,
    source_wb_formulas: Workbook | None = None,
) -> tuple[list[tuple[str, list[object]]], list[tuple[str, list[object]]]] | None:
    formula_resolver = (
        WorkbookFormulaResolver(source_wb_values, source_wb_formulas)
        if source_wb_formulas is not None
        else None
    )
    for resolver in (
        build_legacy_summary_backfill_rows,
        build_legacy_anthro_backfill_rows,
        build_template_backfill_rows,
    ):
        if resolver is build_template_backfill_rows:
            rows = resolver(source_wb_values)
        else:
            rows = resolver(
                source_wb_values,
                formula_resolver=formula_resolver,
            )
        if rows is not None:
            return rows
    return None


def build_template_plan_rows(
    source_wb_values: Workbook,
) -> list[list[object]] | None:
    if PLAN_TEMPLATE_SHEET not in source_wb_values.sheetnames:
        return None
    rows = extract_table_rows(
        source_wb_values[PLAN_TEMPLATE_SHEET],
        canonical_headers=PLAN_TEMPLATE_HEADERS,
        key_header="COMIDA",
    )
    return rows or None


def build_legacy_plan_rows(
    source_wb_values: Workbook,
) -> list[list[object]] | None:
    if LEGACY_PLAN_SHEET not in source_wb_values.sheetnames:
        return None

    ws = source_wb_values[LEGACY_PLAN_SHEET]
    group_headers = [
        ("LACTEOS", "L"),
        ("VEGETALES", "V"),
        ("FRUTAS", "F"),
        ("ALMIDONES", "A"),
        ("PROTEINAS", "P"),
        ("GRASAS", "G"),
    ]

    rows: list[list[object]] = []
    has_non_zero = False
    for meal_def in MEAL_DEFS:
        row = [meal_def["name"]]
        meal_col_idx = column_index_from_string(meal_def["col"])
        for _, group_code in group_headers:
            value = ws.cell(row=GROUP_ROWS[group_code], column=meal_col_idx).value
            parsed = parse_backfill_number(value)
            normalized_value = 0 if parsed is None else parsed
            if normalized_value > 0:
                has_non_zero = True
            row.append(int(normalized_value) if float(normalized_value).is_integer() else normalized_value)
        rows.append(row)

    if not has_non_zero:
        return None
    return rows


def resolve_plan_backfill_rows(
    source_wb_values: Workbook,
) -> list[list[object]] | None:
    for resolver in (
        build_template_plan_rows,
        build_legacy_plan_rows,
    ):
        rows = resolver(source_wb_values)
        if rows is not None:
            return rows
    return None


def backfill_plan_template_sheet(
    target_ws: Worksheet,
    source_wb_values: Workbook,
    template_wb: Workbook,
) -> None:
    rows = resolve_plan_backfill_rows(source_wb_values)
    template_rows = build_template_plan_rows(template_wb)
    if rows == template_rows:
        rows = None
    if rows is None:
        clear_sheet_rows(target_ws, start_row=2)
        apply_compact_table_layout(
            target_ws,
            column_widths=PLAN_TEMPLATE_COLUMN_WIDTHS,
            last_visible_col=len(PLAN_TEMPLATE_HEADERS),
        )
        return
    write_table_rows(target_ws, rows)
    apply_compact_table_layout(
        target_ws,
        column_widths=PLAN_TEMPLATE_COLUMN_WIDTHS,
        last_visible_col=len(PLAN_TEMPLATE_HEADERS),
    )


def build_examples_rows(
    source_wb_values: Workbook,
) -> list[list[object]] | None:
    if EXAMPLES_SHEET not in source_wb_values.sheetnames:
        return None
    rows = extract_table_rows(
        source_wb_values[EXAMPLES_SHEET],
        canonical_headers=EXAMPLES_TEMPLATE_HEADERS,
        key_header="COMIDA",
    )
    return rows or None


def clear_example_observation_values(
    rows: Sequence[Sequence[object]],
) -> list[list[object]]:
    normalized_rows: list[list[object]] = []
    for row in rows:
        row_values = list(row)
        if len(row_values) >= 8:
            row_values[7] = None
        normalized_rows.append(row_values)
    return normalized_rows


def backfill_examples_sheet(
    target_ws: Worksheet,
    source_wb_values: Workbook,
    template_wb: Workbook,
) -> None:
    rows = build_examples_rows(source_wb_values)
    template_rows = build_examples_rows(template_wb)
    if rows == template_rows:
        rows = None
    rows_to_write = rows if rows is not None else template_rows
    if rows_to_write is None:
        apply_compact_table_layout(
            target_ws,
            column_widths=EXAMPLES_TEMPLATE_COLUMN_WIDTHS,
            last_visible_col=len(EXAMPLES_TEMPLATE_HEADERS),
        )
        return
    write_table_rows(
        target_ws,
        clear_example_observation_values(rows_to_write),
    )
    apply_compact_table_layout(
        target_ws,
        column_widths=EXAMPLES_TEMPLATE_COLUMN_WIDTHS,
        last_visible_col=len(EXAMPLES_TEMPLATE_HEADERS),
    )


def build_equivalence_rows(
    source_wb_values: Workbook,
) -> list[list[object]] | None:
    if EXAMPLE_EQUIVALENCES_SHEET not in source_wb_values.sheetnames:
        return None
    rows = extract_table_rows(
        source_wb_values[EXAMPLE_EQUIVALENCES_SHEET],
        canonical_headers=EQUIVALENCE_TEMPLATE_HEADERS,
        key_header="CODIGO ALIMENTO",
    )
    return rows or None


def backfill_equivalence_sheet(
    target_ws: Worksheet,
    source_wb_values: Workbook,
    template_wb: Workbook,
) -> None:
    rows = build_equivalence_rows(source_wb_values)
    template_rows = build_equivalence_rows(template_wb)
    if rows is None or rows == template_rows:
        apply_compact_table_layout(
            target_ws,
            column_widths=EQUIVALENCE_TEMPLATE_COLUMN_WIDTHS,
            last_visible_col=len(EQUIVALENCE_TEMPLATE_COLUMN_WIDTHS),
        )
        return
    write_table_rows(target_ws, rows)
    apply_compact_table_layout(
        target_ws,
        column_widths=EQUIVALENCE_TEMPLATE_COLUMN_WIDTHS,
        last_visible_col=len(EQUIVALENCE_TEMPLATE_COLUMN_WIDTHS),
    )


def write_anthro_template_rows(
    target_ws: Worksheet,
    summary_rows: list[tuple[str, list[object]]],
    measurement_rows: list[tuple[str, list[object]]],
) -> None:
    value_count = max(
        [len(values) for _, values in summary_rows + measurement_rows],
        default=1,
    )
    headers = ["SECCION", "ETIQUETA", "VALOR"] + [
        f"CONTROL_{idx}" for idx in range(1, value_count)
    ]

    target_ws.delete_rows(1, target_ws.max_row)
    target_ws.append(headers)
    for label, values in summary_rows:
        row_values = list(values[:value_count])
        row_values.extend([None] * (value_count - len(row_values)))
        target_ws.append(["RESUMEN", label, *row_values])
    for label, values in measurement_rows:
        row_values = list(values[:value_count])
        row_values.extend([None] * (value_count - len(row_values)))
        target_ws.append(["MEDIDAS", label, *row_values])

    target_ws.freeze_panes = "A2"
    target_ws.column_dimensions["A"].width = target_ws.column_dimensions["A"].width or DEFAULT_COLUMN_WIDTH
    target_ws.column_dimensions["B"].width = target_ws.column_dimensions["B"].width or 34.0
    for col_idx in range(3, 3 + value_count):
        target_ws.column_dimensions[get_column_letter(col_idx)].width = DEFAULT_COLUMN_WIDTH


def backfill_anthro_template_sheet(
    target_ws: Worksheet,
    source_wb_values: Workbook,
    template_wb: Workbook,
    source_wb_formulas: Workbook | None = None,
) -> None:
    rows = resolve_anthro_backfill_rows(
        source_wb_values,
        source_wb_formulas=source_wb_formulas,
    )
    template_rows = build_template_backfill_rows(template_wb)
    if anthro_rows_equal(rows, template_rows):
        rows = None
    if rows is None:
        for row_idx in range(2, target_ws.max_row + 1):
            for col_idx in range(3, target_ws.max_column + 1):
                target_ws.cell(row=row_idx, column=col_idx).value = None
        return
    summary_rows, measurement_rows = rows
    write_anthro_template_rows(target_ws, summary_rows, measurement_rows)


def copy_sheet_cell_contents(source_ws: Worksheet, target_ws: Worksheet) -> None:
    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws.cell(
                row=source_cell.row,
                column=source_cell.column,
                value=source_cell.value,
            )
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)
            if source_cell.border:
                target_cell.border = copy(source_cell.border)
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            if source_cell.protection:
                target_cell.protection = copy(source_cell.protection)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)


def copy_sheet_dimensions(source_ws: Worksheet, target_ws: Worksheet) -> None:
    for key, dimension in source_ws.column_dimensions.items():
        target_dimension = copy(dimension)
        target_dimension.worksheet = target_ws
        target_ws.column_dimensions[key] = target_dimension

    for key, dimension in source_ws.row_dimensions.items():
        target_dimension = copy(dimension)
        target_dimension.worksheet = target_ws
        target_ws.row_dimensions[key] = target_dimension


def copy_sheet_layout(source_ws: Worksheet, target_ws: Worksheet) -> None:
    target_ws.sheet_format = copy(source_ws.sheet_format)
    target_ws.sheet_properties = copy(source_ws.sheet_properties)
    target_ws.page_margins = copy(source_ws.page_margins)
    target_ws.page_setup = copy(source_ws.page_setup)
    target_ws.print_options = copy(source_ws.print_options)
    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_state = source_ws.sheet_state
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines

    if source_ws.auto_filter.ref:
        target_ws.auto_filter.ref = source_ws.auto_filter.ref

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    copy_sheet_dimensions(source_ws, target_ws)
    copy_sheet_cell_contents(source_ws, target_ws)


def create_sheet_from_template(
    target_wb: Workbook,
    template_ws: Worksheet,
    *,
    index: int | None = None,
) -> Worksheet:
    target_ws = target_wb.create_sheet(title=template_ws.title, index=index)
    copy_sheet_layout(template_ws, target_ws)
    return target_ws


def apply_template_sheets_to_workbook(
    target_wb: Workbook,
    *,
    workbook_path: Path,
    template_wb: Workbook,
    source_wb_values: Workbook,
    source_wb_formulas: Workbook,
    sheet_names: Sequence[str],
    replace_existing: bool,
) -> WorkbookSheetUpdateResult:
    added_sheets: list[str] = []
    replaced_sheets: list[str] = []
    skipped_sheets: list[str] = []

    for sheet_name in sheet_names:
        if sheet_name not in template_wb.sheetnames:
            raise ValueError(
                f"La hoja {sheet_name} no existe en el workbook template."
            )

        template_ws = template_wb[sheet_name]
        target_index: int | None = None

        if sheet_name in target_wb.sheetnames:
            if not replace_existing:
                skipped_sheets.append(sheet_name)
                continue
            target_index = target_wb.sheetnames.index(sheet_name)
            target_wb.remove(target_wb[sheet_name])
            replaced_sheets.append(sheet_name)

        create_sheet_from_template(target_wb, template_ws, index=target_index)
        if sheet_name == PLAN_TEMPLATE_SHEET:
            backfill_plan_template_sheet(
                target_wb[sheet_name],
                source_wb_values,
                template_wb,
            )
        elif sheet_name == ANTHRO_TEMPLATE_SHEET:
            backfill_anthro_template_sheet(
                target_wb[sheet_name],
                source_wb_values,
                template_wb,
                source_wb_formulas,
            )
        elif sheet_name == EXAMPLES_SHEET:
            backfill_examples_sheet(
                target_wb[sheet_name],
                source_wb_values,
                template_wb,
            )
        elif sheet_name == EXAMPLE_EQUIVALENCES_SHEET:
            backfill_equivalence_sheet(
                target_wb[sheet_name],
                source_wb_values,
                template_wb,
            )
        if target_index is None:
            added_sheets.append(sheet_name)

    if added_sheets or replaced_sheets:
        target_wb.save(workbook_path)

    return WorkbookSheetUpdateResult(
        path=workbook_path,
        added_sheets=tuple(added_sheets),
        replaced_sheets=tuple(replaced_sheets),
        skipped_sheets=tuple(skipped_sheets),
    )


def update_workbook_with_template_sheets(
    workbook_path: Path | str,
    *,
    template_workbook_path: Path | str = DEFAULT_TEMPLATE_WORKBOOK,
    sheet_names: Sequence[str] = DEFAULT_SHEETS_TO_ADD,
    replace_existing: bool = True,
) -> WorkbookSheetUpdateResult:
    workbook_path = Path(workbook_path)
    template_workbook_path = Path(template_workbook_path)

    if workbook_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Formato no soportado: {workbook_path.name}")
    if not workbook_path.exists():
        raise FileNotFoundError(f"No existe el archivo: {workbook_path}")
    if not template_workbook_path.exists():
        raise FileNotFoundError(
            f"No existe el workbook template: {template_workbook_path}"
        )

    template_wb = load_excel_workbook(template_workbook_path)
    target_wb = load_excel_workbook(workbook_path)
    source_wb_values = load_excel_workbook(workbook_path, data_only=True)
    source_wb_formulas = load_excel_workbook(workbook_path)
    try:
        return apply_template_sheets_to_workbook(
            target_wb,
            workbook_path=workbook_path,
            template_wb=template_wb,
            source_wb_values=source_wb_values,
            source_wb_formulas=source_wb_formulas,
            sheet_names=sheet_names,
            replace_existing=replace_existing,
        )
    finally:
        source_wb_formulas.close()
        source_wb_values.close()
        target_wb.close()
        template_wb.close()


def update_folder_with_template_sheets(
    folder: Path | str,
    *,
    template_workbook_path: Path | str = DEFAULT_TEMPLATE_WORKBOOK,
    sheet_names: Sequence[str] = DEFAULT_SHEETS_TO_ADD,
    replace_existing: bool = True,
    recursive: bool = True,
) -> list[WorkbookSheetUpdateResult]:
    folder = Path(folder)
    template_workbook_path = Path(template_workbook_path)
    if not folder.exists():
        raise FileNotFoundError(f"No existe la carpeta: {folder}")
    if not folder.is_dir():
        raise NotADirectoryError(f"La ruta no es una carpeta: {folder}")
    if not template_workbook_path.exists():
        raise FileNotFoundError(
            f"No existe el workbook template: {template_workbook_path}"
        )

    template_wb = load_excel_workbook(template_workbook_path)
    results: list[WorkbookSheetUpdateResult] = []
    template_resolved = template_workbook_path.resolve()
    try:
        for workbook_path in iter_excel_files(folder, recursive=recursive):
            if workbook_path.resolve() == template_resolved:
                continue

            target_wb = load_excel_workbook(workbook_path)
            source_wb_values = load_excel_workbook(workbook_path, data_only=True)
            source_wb_formulas = load_excel_workbook(workbook_path)
            try:
                result = apply_template_sheets_to_workbook(
                    target_wb,
                    workbook_path=workbook_path,
                    template_wb=template_wb,
                    source_wb_values=source_wb_values,
                    source_wb_formulas=source_wb_formulas,
                    sheet_names=sheet_names,
                    replace_existing=replace_existing,
                )
            finally:
                source_wb_formulas.close()
                source_wb_values.close()
                target_wb.close()
            results.append(result)
        return results
    finally:
        template_wb.close()


def build_result_line(result: WorkbookSheetUpdateResult) -> str:
    if not result.changed:
        skipped = ", ".join(result.skipped_sheets) or "sin cambios"
        return f"= {result.path.name}: {skipped}"

    parts: list[str] = []
    if result.added_sheets:
        parts.append("agregadas=" + ", ".join(result.added_sheets))
    if result.replaced_sheets:
        parts.append("reemplazadas=" + ", ".join(result.replaced_sheets))
    if result.skipped_sheets:
        parts.append("omitidas=" + ", ".join(result.skipped_sheets))
    return f"+ {result.path.name}: " + " | ".join(parts)


def main() -> int:
    args = parse_args()
    folder = Path(args.folder)
    template_workbook_path = DEFAULT_TEMPLATE_WORKBOOK
    sheet_names = selected_sheet_names()

    try:
        results = update_folder_with_template_sheets(
            folder,
            template_workbook_path=template_workbook_path,
            sheet_names=sheet_names,
        )
    except (FileNotFoundError, NotADirectoryError, ValueError) as exc:
        print(f"Error: {exc}")
        return 1

    if not results:
        print("No se encontraron archivos .xlsx o .xlsm para procesar.")
        return 1

    changed_count = sum(1 for result in results if result.changed)
    for result in results:
        print(build_result_line(result))

    print(
        f"Procesados: {len(results)} archivo(s). "
        f"Modificados: {changed_count}. "
        f"Sin cambios: {len(results) - changed_count}."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
